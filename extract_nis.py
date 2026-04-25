"""Extract NIS curriculum data from the 4 Excel workbooks."""
import openpyxl
import os
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

BASE = r"C:\Users\Paolo\OneDrive\Documentos\Claude\Projects_I\Information"

FILES = {
    "grammar": "NIS_Grammar_Curriculum_Complete.xlsx",
    "ela": "NIS_ELA_CURRICULUM_COMPLETE.xlsx",
    "units": "OK_Units_by_Grade_G1-G11_Nordic.V2.xlsx",
    "ss": "S&S English.xlsx",
}

def dump_sheet(ws, max_rows=120, max_cols=15):
    rows = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx > max_rows:
            rows.append(["...(truncated)..."])
            break
        trimmed = [("" if v is None else str(v)) for v in row[:max_cols]]
        # drop trailing empties
        while trimmed and trimmed[-1] == "":
            trimmed.pop()
        if trimmed:
            rows.append(trimmed)
    return rows

def dump_workbook(path, label):
    print(f"\n{'='*100}\n### WORKBOOK: {label} -> {os.path.basename(path)}\n{'='*100}")
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n--- SHEET: {sn}  (dim={ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column}) ---")
        rows = dump_sheet(ws)
        for r in rows:
            print(" | ".join(r))

for k, fn in FILES.items():
    p = os.path.join(BASE, fn)
    if os.path.exists(p):
        dump_workbook(p, k)
    else:
        print(f"MISSING: {p}")
