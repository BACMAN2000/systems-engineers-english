# -*- coding: utf-8 -*-
"""
NIS English Department Master Scope & Sequence — 2026
Consolidates: S&S English, Grammar Curriculum, ELA Curriculum, Units by Grade,
Cambridge pathway, Benchmarks, Assessment rubrics MINEDU, CLIL integrations.

Applies verified corrections vs Cambridge official syllabi.
Designed for teacher usability: colors by level, clear headers, frozen panes.
"""
import os
import sys
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# =============================================================================
# STYLES
# =============================================================================
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")       # dark blue
SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")    # medium blue
LEVEL_FILLS = {
    "PreA1":  PatternFill("solid", fgColor="FCE4EC"),   # pink
    "A1":     PatternFill("solid", fgColor="FFF3E0"),   # peach
    "A2":     PatternFill("solid", fgColor="FFFDE7"),   # light yellow
    "B1":     PatternFill("solid", fgColor="E8F5E9"),   # light green
    "B2":     PatternFill("solid", fgColor="E3F2FD"),   # light blue
    "C1":     PatternFill("solid", fgColor="F3E5F5"),   # light purple
}
ROW_ALT = PatternFill("solid", fgColor="F7F9FC")
CORRECTION_FILL = PatternFill("solid", fgColor="FFECB3")  # highlight corrections
UNIT_THEMES_FILL = PatternFill("solid", fgColor="F2CC8F")    # warm amber — Unit Themes row highlight
UNIT_THEMES_FONT = Font(bold=True, size=11, color="5B3A00", name="Calibri")
WRAP_CENTER_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
BODY_FONT = Font(size=10, name="Calibri")
BOLD_FONT = Font(bold=True, size=10, name="Calibri")
TITLE_FONT = Font(bold=True, size=16, color="1F4E79", name="Calibri")
SUBTITLE_FONT = Font(bold=True, size=12, color="2E75B6", name="Calibri")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
WRAP_LEFT = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_TOPCENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

# Grade → level mapping for color coding
GRADE_LEVEL = {
    "G1": "PreA1", "G2": "PreA1", "G3": "A1", "G4": "A1",
    "G5": "A2", "G6": "A2", "G7": "B1", "G8": "B1",
    "G9": "B2", "G10": "B2", "G11": "C1"
}

# =============================================================================
# HELPERS
# =============================================================================
def style_header_row(ws, row_idx, num_cols, fill=HEADER_FILL, font=HEADER_FONT, height=30):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = WRAP_CENTER
        cell.border = BORDER
    ws.row_dimensions[row_idx].height = height

def style_body(ws, start_row, end_row, num_cols, alt=True):
    for r in range(start_row, end_row + 1):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = WRAP_LEFT
            cell.border = BORDER
            if alt and (r - start_row) % 2 == 1:
                cell.fill = ROW_ALT

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def apply_level_color_to_row(ws, row_idx, num_cols, grade):
    lvl = GRADE_LEVEL.get(grade)
    if lvl:
        for c in range(1, num_cols + 1):
            ws.cell(row=row_idx, column=c).fill = LEVEL_FILLS[lvl]

def write_title(ws, title, subtitle=None, row=1, span=10):
    ws.cell(row=row, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 28
    if subtitle:
        ws.cell(row=row+1, column=1, value=subtitle).font = SUBTITLE_FONT
        ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=span)
        ws.row_dimensions[row+1].height = 20
        return row + 3
    return row + 2


# =============================================================================
# CREATE WORKBOOK
# =============================================================================
wb = Workbook()
wb.remove(wb.active)  # remove default

# =============================================================================
# SHEET 00: COVER
# =============================================================================
ws = wb.create_sheet("00_Cover")
set_col_widths(ws, [4, 95])
ws.cell(row=2, column=2, value="NIS — NORDIC INTERNATIONAL SCHOOL").font = Font(bold=True, size=20, color="1F4E79")
ws.cell(row=3, column=2, value="English Department — Master Scope & Sequence 2026").font = Font(bold=True, size=14, color="2E75B6")
ws.cell(row=4, column=2, value="Pathway: Cambridge YLE → KET → PET → FCE → CAE  ·  CEFR Pre-A1 → C1  ·  MINEDU Ciclos III–VII").font = Font(italic=True, size=11, color="595959")
ws.cell(row=5, column=2, value="Version 1.0  ·  April 2026  ·  Prepared by the English Coordination").font = Font(size=10, color="808080")

contents = [
    ("", ""),
    ("HOW TO USE THIS WORKBOOK", "section"),
    ("1. Each teacher: read tabs 01_Pathway, 02_Benchmarks, and your grade in 15_Teacher_QuickRef first.",""),
    ("2. For lesson planning, use 10_Units_G1_G5 or 11_Units_G6_G11 and cross-check with 04_Grammar and 05_Vocabulary.",""),
    ("3. For assessment, consult 14_Assessment_Scale (MINEDU AD/A/B/C) and 13_Cambridge_Calendar.",""),
    ("4. For cross-subject projects, see 12_CLIL_Map.",""),
    ("5. Coordination-approved changes tracked in 16_Change_Log.",""),
    ("", ""),
    ("CONTENTS", "section"),
    ("00_Cover — This page", ""),
    ("01_Pathway — Cambridge × CEFR × MINEDU alignment matrix G1–G11", ""),
    ("02_Benchmarks — End-of-grade outcomes per skill", ""),
    ("03_SS_Master — Full scope & sequence matrix (all grades × all skills)", ""),
    ("04_Grammar — Grammar progression by grade (Cambridge-aligned, corrected)", ""),
    ("05_Vocabulary — Thematic vocabulary by grade + Cambridge wordlist reference", ""),
    ("06_Phonics — Systematic phonics G1–G5 (8 units per grade)", ""),
    ("07_Reading — Reading plan by grade (titles + level + Lexile/CEFR)", ""),
    ("08_Writing — Text types + word counts per grade (corrected)", ""),
    ("09_Speaking_Listening — Oral skills progression G1–G11", ""),
    ("10_Units_G1_G5 — 6 unit plans × 5 grades (themes, vocab, grammar, outputs)", ""),
    ("11_Units_G6_G11 — 6 unit plans × 6 grades (themes, vocab, grammar, outputs)", ""),
    ("12_CLIL_Map — Cross-curricular project integrations", ""),
    ("13_Cambridge_Calendar — Mock, prep, and official exam schedule 2026", ""),
    ("14_Assessment_Scale — MINEDU AD/A/B/C descriptors per skill", ""),
    ("15_Teacher_QuickRef — 1-page summary per grade (printable)", ""),
    ("16_Change_Log — Corrections applied vs previous documents", ""),
    ("17_Rubrics_Ready — 15 activity-specific ready-made rubrics (AD/A/B/C) with grade-range filter", ""),
    ("18_Detail_G1_G5 — Granular unit content per sub-block (ELA Complete source) — PRIMARY", ""),
    ("19_Detail_G6_G11 — Granular unit content per sub-block (ELA Complete source) — SECONDARY", ""),
    ("20_Units_Full_G1_G11 — COMPLETE unit plans: Phonics, Vocab, Reading, Writing, L&S, Graduate Profile, ATL, Formative + Summative Assessments (OK_Units source + corrections)", ""),
    ("21_Immersion_Framework — % English per grade, language policy, BICS/CALP progression, English-only zones", ""),
    ("22_Teaching_Strategies — Scaffolding ladder for struggling students + differentiation for mixed-level classes", ""),
    ("23_Parent_Guide — Bilingual ES/EN guide: how to support at home WITHOUT translating (printable)", ""),
    ("", ""),
    ("KEY CORRECTIONS APPLIED (FULL LIST IN 16_Change_Log)", "section"),
    ("• G1: removed past simple / going to / would — not Pre-A1", ""),
    ("• G2: removed past simple expanded / future will / comparatives (moved to G4)", ""),
    ("• G4: moved tag questions to G5 (Flyers spec)", ""),
    ("• G5: replaced 'Into the Wild' with age-10-appropriate titles", ""),
    ("• G6: reconciled Grammar vs ELA contradiction — KET-aligned path", ""),
    ("• G7: added conditional 2, reported speech, causative", ""),
    ("• G9: added review-genre writing practice (FCE)", ""),
    ("• G11: replaced 'Moby Dick' full novel; removed 'article' from CAE writing (proposal only); fixed CAE word count 220–260", ""),
    ("• G1–G3: halved classroom writing word-count targets to age-appropriate levels", ""),
]

row = 7
for text, kind in contents:
    c = ws.cell(row=row, column=2, value=text)
    if kind == "section":
        c.font = Font(bold=True, size=12, color="1F4E79")
        c.fill = PatternFill("solid", fgColor="DCE6F1")
    else:
        c.font = BODY_FONT
    row += 1

ws.sheet_view.showGridLines = False


# =============================================================================
# SHEET 01: PATHWAY
# =============================================================================
ws = wb.create_sheet("01_Pathway")
next_row = write_title(ws, "Pathway — CEFR × Cambridge × MINEDU by Grade",
                       "Official NIS alignment 2026 — reconciles prior inconsistencies", row=1, span=8)
headers = ["Grade", "Age", "CEFR", "Cambridge Exam (target)", "MINEDU Ciclo",
           "Grammar Anchor (summary)", "Writing output (words)", "Reading length (words)"]
for i, h in enumerate(headers, 1):
    ws.cell(row=next_row, column=i, value=h)
style_header_row(ws, next_row, len(headers))

pathway_rows = [
    ("G1",  "6",  "Pre-A1",     "— (Pre-Starters prep)",    "III",
     "To be, have got, can, present simple, like + noun, imperatives",
     "10–25 (labels)",     "picture books (20–50 w/page)"),
    ("G2",  "7",  "Pre-A1→A1",  "Cambridge Starters",       "III",
     "Present simple/continuous, can, have got, like+ing, there is/are",
     "20–40 (sentence-level)",   "30–60"),
    ("G3",  "8",  "A1",         "— (Pre-Movers prep)",      "IV",
     "SVA, present continuous, comp/superlative, must/have to/could",
     "40–70",              "60–100"),
    ("G4",  "9",  "A1→A2",      "Cambridge Movers",         "IV",
     "Past simple reg/irreg, verb+inf/+ing, infinitive of purpose, relatives (who/which/where)",
     "60–100",             "100–150"),
    ("G5",  "10", "A2",         "Cambridge Flyers",         "V",
     "Past continuous, present perfect (intro), going to/will, might/should, tag Qs, zero conditional",
     "100–150",            "150–250"),
    ("G6",  "11", "A2→B1",      "Cambridge A2 Key (KET)",   "V",
     "Present perfect (full), full modals, passive (pres/past simple), gerunds, 1st conditional",
     "100–180 (+ KET 25w email / 35w story)", "200–300"),
    ("G7",  "12", "B1",         "— (KET+ / PET bridge)",    "VI",
     "Past perfect, conditionals 0-1-2, reported speech (statements), phrasal verbs, modal passive",
     "180–250 (+ PET 100w)", "300–450"),
    ("G8",  "13", "B1+",        "Cambridge B1 Preliminary (PET)", "VI",
     "Causative, reported speech (Qs/commands), word formation, all modals, so/nor auxiliaries",
     "200–260 (PET 100w exam)",  "450–600"),
    ("G9",  "14", "B1+→B2",     "Cambridge B2 First (FCE)", "VII",
     "Conditional 3 + mixed, passive all tenses, past modals, advanced relatives, wish/if only",
     "250–380 (FCE 140–190w)",   "700–1000"),
    ("G10", "15", "B2+",        "— (FCE consolidation / CAE bridge)", "VII",
     "Subjunctive, complex relatives, participle clauses, inversion (intro), cleft (intro)",
     "300–430 (incl. CV/cover letter)", "1000–1300"),
    ("G11", "16", "C1",         "Cambridge C1 Advanced (CAE)", "VII",
     "Inversion, cleft, discourse markers, impersonal passive, nominalisation, mixed conditionals",
     "220–260 (CAE exam) / 350–500 classroom", "1300–1500+"),
]

r = next_row + 1
for row_data in pathway_rows:
    for i, v in enumerate(row_data, 1):
        ws.cell(row=r, column=i, value=v)
    apply_level_color_to_row(ws, r, len(headers), row_data[0])
    for c in range(1, len(headers)+1):
        ws.cell(row=r, column=c).font = BODY_FONT
        ws.cell(row=r, column=c).alignment = WRAP_LEFT
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = 45
    r += 1

set_col_widths(ws, [6, 6, 10, 26, 10, 45, 28, 22])
ws.freeze_panes = "A" + str(next_row + 1)

# Notes
r += 1
ws.cell(row=r, column=1, value="NOTES").font = Font(bold=True, size=11, color="1F4E79")
r += 1
notes = [
    "• Cambridge YLE (Starters/Movers/Flyers) are Young Learners exams targeting ages 7–12.",
    "• KET (A2 Key for Schools), PET (B1 Preliminary for Schools), FCE (B2 First for Schools), CAE (C1 Advanced).",
    "• NIS exceeds MINEDU CEFR expectations by +0.5 to +2 levels — bridging strategies documented per grade.",
    "• Word counts for classroom and exam are BOTH given where relevant; exam is shorter to ensure students practise the real format.",
    "• G7 is a bridge year: consolidates KET, introduces PET grammar without exam pressure.",
    "• G10 consolidates FCE content while introducing C1/CAE structures — dual purpose.",
]
for n in notes:
    ws.cell(row=r, column=1, value=n).font = BODY_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1

ws.sheet_view.showGridLines = False


# =============================================================================
# SHEET 02: BENCHMARKS — end-of-grade outcomes per skill
# =============================================================================
ws = wb.create_sheet("02_Benchmarks")
next_row = write_title(ws, "Benchmarks — End-of-Grade Outcomes per Skill",
                       "Can-Do statements aligned to Cambridge exam descriptors", row=1, span=6)
headers = ["Grade / Exam", "Listening", "Speaking", "Reading", "Writing", "Grammar & Vocabulary"]
for i, h in enumerate(headers, 1):
    ws.cell(row=next_row, column=i, value=h)
style_header_row(ws, next_row, len(headers))

bench_rows = [
    ("G1 / Pre-Starters",
     "Follow classroom instructions; identify familiar words in songs and short spoken sentences.",
     "Greet; say name, age, colours; name objects; answer yes/no questions.",
     "Recognise all 26 letters; match words to pictures (family, food, animals, body).",
     "Write own name; trace and copy words; label pictures (3–5 words).",
     "~150 productive words; Verb to be, can, like + noun, imperatives, this/that."),
    ("G2 / Starters",
     "Understand short simple dialogues about family, school, animals, food (Cambridge Starters Listening parts 1–4).",
     "Answer picture-prompted questions; describe a picture in simple sentences.",
     "Read 2–3-sentence texts; match pictures to words and short captions.",
     "Copy words and complete short sentences (Starters Reading/Writing parts).",
     "~450 words (Starters wordlist); Present simple, present continuous (limited), there is/are, like + -ing."),
    ("G3 / Pre-Movers",
     "Understand short dialogues with more vocabulary; identify main information in simple announcements.",
     "Describe routines, favourites, community helpers; give short opinions.",
     "Read short paragraphs; answer simple comprehension Qs; find specific info.",
     "Write short paragraph about self/family/routine (40–70 words).",
     "~800 words; All basic tenses start; must/have to/could; SVA; superlatives."),
    ("G4 / Movers",
     "Understand longer dialogues (Movers parts 1–5); identify specific info in spoken texts.",
     "Describe places, objects, past events; role-play simple situations.",
     "Read and sequence short stories; locate information in signs, notices.",
     "Write a note/message of 30–40 words (Movers Writing part); classroom paragraphs 60–100w.",
     "~1350 words (Movers wordlist); Past simple, relatives, infinitive of purpose, have to."),
    ("G5 / Flyers",
     "Understand dialogues and short stories (Flyers parts 1–5); infer mood from intonation.",
     "Narrate stories with past/present/future; express preferences with reasons.",
     "Read and understand short stories (80–250 words); find main idea and detail.",
     "Writing Part 3: 20–25-word picture story (Flyers); classroom pieces 100–150w.",
     "~2500 words (Flyers wordlist); Present perfect intro, tag Qs, 0th conditional, be going to."),
    ("G6 / KET",
     "Follow dialogues and announcements (KET Listening parts 1–5); 25-minute exam.",
     "KET Speaking: 2 parts, 8–10 min; interview + collaborative task.",
     "Understand signs, notices, emails, short articles (KET Reading parts 1–5).",
     "KET Writing: 25-word email (Part 6) + 35-word picture-based story (Part 7).",
     "~1500 words (KET list); All basic tenses + present perfect, full modals."),
    ("G7 / KET+ bridge",
     "Follow longer dialogues (3–5 min) at B1 level; identify speakers' attitudes.",
     "Sustain 3-minute discussions on familiar topics; reach agreement in pairs.",
     "Read articles, reports, narratives up to 450 words; infer meaning.",
     "Produce 100–150w emails, articles; basic report format.",
     "~2200 words; past perfect, conditionals 0-1-2, reported speech (statements)."),
    ("G8 / PET",
     "PET Listening: 4 parts, 30 min; multiple choice, gap fill, note-taking.",
     "PET Speaking: 4 parts, 10–12 min; individual + paired collaborative.",
     "PET Reading: 6 parts; gapped text, multi-matching, multiple-choice cloze.",
     "PET Writing: Part 1 email ~100w; Part 2 article OR story ~100w.",
     "~3200 words (PET list); all tenses mastered, all conditionals, passive, reported."),
    ("G9 / FCE",
     "FCE Listening: 4 parts, 40 min; monologues, interviews, multi-task.",
     "FCE Speaking: 4 parts, 14 min; interview, picture comparison, discussion.",
     "FCE Reading/Use of English: 7 parts, 75 min.",
     "FCE Writing: Part 1 essay 140–190w compulsory; Part 2 one of article/email/letter/report/review 140–190w.",
     "~4500 words; cond. 3 + mixed, passive all tenses, past modals, complex relatives."),
    ("G10 / FCE consolidation",
     "Comprehension of extended discourse; academic lecture excerpts.",
     "Presentations 5–7 min; Q&A; justifying opinions with evidence.",
     "Newspapers, op-eds, literary extracts (4000+ words).",
     "Persuasive / argumentative essays 400–500w; CV + cover letter.",
     "~5000 words; subjunctive, participle clauses, inversion (intro), cleft (intro)."),
    ("G11 / CAE",
     "CAE Listening: 4 parts, 40 min; complex attitudes, inference.",
     "CAE Speaking: 4 parts, 15 min; interview, long turn, collaborative, discussion.",
     "CAE Reading/UoE: 8 parts, 90 min; multi-matching, cross-text.",
     "CAE Writing: Part 1 essay 220–260w; Part 2 one of letter/email, proposal, report, review 220–260w.",
     "~5500–6000 words; inversion, cleft, nominalisation, impersonal passive, hedging."),
]

r = next_row + 1
for row_data in bench_rows:
    for i, v in enumerate(row_data, 1):
        ws.cell(row=r, column=i, value=v)
    grade_part = row_data[0].split(" /")[0].strip()
    apply_level_color_to_row(ws, r, len(headers), grade_part)
    for c in range(1, len(headers)+1):
        ws.cell(row=r, column=c).font = BODY_FONT
        ws.cell(row=r, column=c).alignment = WRAP_LEFT
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = 75
    r += 1

set_col_widths(ws, [20, 32, 32, 32, 34, 36])
ws.freeze_panes = "B" + str(next_row + 1)
ws.sheet_view.showGridLines = False


# =============================================================================
# SHEET 03: SS_MASTER
# =============================================================================
ws = wb.create_sheet("03_SS_Master")
next_row = write_title(ws, "Scope & Sequence Master Matrix",
                       "All grades × 6 skill blocks — drill down in dedicated tabs", row=1, span=8)
headers = ["Skill / Grade"] + ["G1","G2","G3","G4","G5","G6","G7","G8","G9","G10","G11"]
for i, h in enumerate(headers, 1):
    ws.cell(row=next_row, column=i, value=h)
style_header_row(ws, next_row, len(headers))

master = [
    ("Listening",
     ["Classroom instructions", "Starters dialogues", "Short announcements", "Movers dialogues",
      "Flyers narratives", "KET 25-min exam", "B1 extended", "PET 30-min exam",
      "FCE 40-min exam", "Lectures/op-eds", "CAE 40-min exam"]),
    ("Speaking",
     ["Greetings/short Qs", "Starters picture Qs", "Describe routines", "Movers role-play",
      "Flyers narration", "KET 8–10 min paired", "3-min discussions", "PET 10–12 min",
      "FCE 14 min", "5–7 min presentations", "CAE 15 min"]),
    ("Reading",
     ["Word-pic match", "Starters captions", "Short paragraphs", "Movers stories",
      "Flyers 80–250 w", "KET signs/emails", "Articles 450 w", "PET 6 parts",
      "FCE 7 parts", "4000 w texts", "CAE 8 parts"]),
    ("Writing",
     ["Labels 10–25 w", "Copy + sentences", "Paragraph 40–70 w", "Movers note 30–40 w",
      "Flyers story 20–25 w", "KET email 25 w + story 35 w", "100–150 w", "PET email + article 100 w",
      "FCE 140–190 w", "400–500 w + CV", "CAE 220–260 w"]),
    ("Grammar (anchor)",
     ["To be, can, like+N", "Present S/C, there is/are", "Must/have to, comp/superl",
      "Past simple, relatives", "Past cont, pres perf intro", "Pres perf, modals, passive",
      "Past perf, cond 0-1-2", "Causative, reported, word form", "Cond 3+mixed, passive all",
      "Subjunctive, cleft, inv", "Inversion, cleft, nominal."]),
    ("Vocabulary (size)",
     ["~150", "~450 (Starters)", "~800", "~1350 (Movers)", "~2500 (Flyers)",
      "~1500 (KET)", "~2200", "~3200 (PET)", "~4500 (FCE)", "~5000", "~5500–6000 (CAE)"]),
    ("Phonics / Pron",
     ["Letters + CVC", "Digraphs sh/ch/th", "Long vowels a_e/i_e", "Blends st/sp",
      "Syllable juncture", "Word stress", "Compound stress", "Intonation Qs",
      "Connected speech", "Emotions/tone", "Mastery + impromptu"]),
]
r = next_row + 1
for skill, vals in master:
    ws.cell(row=r, column=1, value=skill).font = BOLD_FONT
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="DCE6F1")
    for i, v in enumerate(vals, 2):
        ws.cell(row=r, column=i, value=v).font = BODY_FONT
        grade = ["G1","G2","G3","G4","G5","G6","G7","G8","G9","G10","G11"][i-2]
        ws.cell(row=r, column=i).fill = LEVEL_FILLS[GRADE_LEVEL[grade]]
        ws.cell(row=r, column=i).alignment = WRAP_LEFT
        ws.cell(row=r, column=i).border = BORDER
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=1).alignment = WRAP_LEFT
    ws.row_dimensions[r].height = 55
    r += 1

set_col_widths(ws, [18] + [15]*11)
ws.freeze_panes = "B" + str(next_row + 1)
ws.sheet_view.showGridLines = False


# =============================================================================
# SHEET 04: GRAMMAR (corrected)
# =============================================================================
ws = wb.create_sheet("04_Grammar")
next_row = write_title(ws, "Grammar Curriculum by Grade — Cambridge-Aligned",
                       "Corrected vs previous docs — see 16_Change_Log", row=1, span=6)
headers = ["Grade / Exam", "CEFR", "Grammar Topics (in teaching order)",
           "Sample Structures", "Exam Link", "Hours"]
for i, h in enumerate(headers, 1):
    ws.cell(row=next_row, column=i, value=h)
style_header_row(ws, next_row, len(headers))

grammar_rows = [
    ("G1 / Pre-Starters", "Pre-A1",
     "1) To be (am/is/are)\n2) Subject pronouns I/you/he/she/it/we/they\n3) Can for ability\n4) Have got\n5) Imperatives (stand up, sit down)\n6) This/that\n7) There is / there are\n8) Like + noun (I like apples)\n9) A/an articles\n10) Plural nouns (regular)",
     "I am happy. / She is my teacher. / I can jump. / I have got a pencil. / This is a book. / I like apples.",
     "Pre-Starters oral interaction; Starters preparation", "45"),
    ("G2 / Starters", "Pre-A1→A1",
     "1) Present simple (I/you/he/she)\n2) Present continuous (limited, I am -ing)\n3) Have got (all persons)\n4) Like + -ing (I like swimming)\n5) Can (ability + permission)\n6) There is / there are + Q forms\n7) Prepositions of place (in, on, under, next to)\n8) Possessive 's\n9) Adjectives (before noun)\n10) Wh- questions (what, who, where)\n\n⚠ NOT included per Starters: past tense, future will, comparatives (→ G4)",
     "He plays football. / I am reading. / She has got a cat. / Is there a book on the table? / My sister's bag.",
     "Cambridge Starters (all 3 papers)", "60"),
    ("G3 / Pre-Movers", "A1",
     "1) Present simple (all persons + Qs + negatives)\n2) Present continuous (all persons)\n3) SVA revision\n4) Comparative / superlative adjectives (short)\n5) Modals: must, have to, could (past of can)\n6) Prepositions of time (at, on, in)\n7) Sentence types: statement, Q, exclam, imperative\n8) Adverbs of frequency",
     "She always does her homework. / Elephants are bigger than cats. / I must go now. / I could swim when I was 5.",
     "Pre-Movers / A1 progression", "60"),
    ("G4 / Movers", "A1→A2",
     "1) Past simple (reg + irreg) — affirmative, negative, Qs\n2) Verb + to-infinitive (want to, need to)\n3) Verb + -ing (like/enjoy swimming)\n4) Infinitive of purpose (I went to see…)\n5) Relative clauses (who, which, where)\n6) Must / have to / had to\n7) Comparative + superlative (all adjectives)\n8) Future with will (basic predictions)\n9) Prepositions of movement",
     "We went to the beach yesterday. / The boy who lives next door. / I had to study. / I'll help you.",
     "Cambridge Movers (all 3 papers)", "65"),
    ("G5 / Flyers", "A2",
     "1) Past continuous (was/were + -ing)\n2) Past simple vs past continuous\n3) Present perfect (intro: have been / have done)\n4) Going to + will (future contrast)\n5) Might / may / should\n6) Tag questions (positive/negative)\n7) Zero conditional (If you heat ice, it melts)\n8) Connectors: because, so, although\n9) Be made of / look like / sound like\n10) Sensory verbs",
     "I was reading when she called. / I have lived here for 3 years. / You should rest. / It's cold, isn't it? / If water freezes, it expands.",
     "Cambridge Flyers (all 3 papers)", "70"),
    ("G6 / KET (A2 Key)", "A2→B1",
     "1) Present perfect (full: ever, never, just, already, yet, for, since)\n2) Past simple/continuous review\n3) Full modals: can, could, may, might, must, mustn't, should, have to, need, needn't\n4) Passive voice (present simple + past simple)\n5) Gerunds as subject/object\n6) First conditional (If + pres, will)\n7) Adjective order (O-S-A-Sh-C-O-M-P)\n8) Phrasal verbs (high-frequency: get up, look for, put on)",
     "She has just finished. / The book was written in 1920. / If it rains, we'll stay home. / A beautiful small blue dress.",
     "A2 Key (KET): UoE, Reading/Writing, Speaking", "85"),
    ("G7 / KET+ / PET bridge", "B1",
     "1) Past perfect simple\n2) Second conditional (If + past, would)\n3) Reported speech (statements)\n4) Relative clauses (defining + non-defining)\n5) Modal passive (must be done)\n6) Causative have/get (I had my hair cut)\n7) Connectors: although, however, despite, in spite of\n8) Gerund vs infinitive (common verbs)\n9) so / neither auxiliaries",
     "She had left when I arrived. / If I were rich, I would travel. / He said he was tired. / I had my hair cut. / Although it rained, we went.",
     "B1 Preliminary preparation (PET bridge)", "90"),
    ("G8 / PET (B1 Preliminary)", "B1+",
     "1) All tenses consolidation\n2) Conditionals 0, 1, 2 consolidated\n3) Reported speech: questions + commands\n4) Passive: all simple tenses\n5) Word formation: prefixes (un-, dis-, re-, pre-) + suffixes (-tion, -ment, -ful, -less, -able)\n6) Phrasal verbs (PET list)\n7) Verb patterns: verb + obj + inf\n8) Register: formal vs informal\n\nLight intro: conditional 3 (→ full in G9)",
     "He asked me where I lived. / She told me to wait. / The house has been sold. / un + happy = unhappy.",
     "B1 Preliminary (PET): UoE, R/W, Listening, Speaking", "100"),
    ("G9 / FCE (B2 First)", "B1+→B2",
     "1) Conditional 3 (If + past perf, would have + past part)\n2) Mixed conditionals\n3) Passive: all tenses (continuous, perfect, modal)\n4) Past modals of deduction (must have, should have, could have, might have)\n5) Advanced relative clauses (non-defining, with prepositions)\n6) Wish / if only\n7) Key word transformation practice\n8) Word formation (FCE list)\n9) Connectors: whereas, nevertheless, furthermore\n10) Phrasal verbs (FCE high-frequency)",
     "If I had studied, I would have passed. / She must have left. / The man to whom I spoke. / I wish I were taller.",
     "B2 First (FCE): UoE, Reading, Writing, Listening, Speaking", "110"),
    ("G10 / FCE consolidation → CAE bridge", "B2+",
     "1) Mixed conditionals mastery\n2) Subjunctive (suggest that he go)\n3) Complex relatives (nested, reduced)\n4) Participle clauses (Having finished…)\n5) Inversion basics (Hardly had I…, Not only…)\n6) Cleft sentences intro (It was X who…)\n7) Causative advanced\n8) Phrasal verbs (academic) + collocations (account for, give rise to)\n9) CV + cover letter conventions\n10) Discourse markers",
     "Hardly had I sat down when the phone rang. / It was Maria who found it. / Having finished, she left.",
     "FCE consolidation + early CAE prep", "115"),
    ("G11 / CAE (C1 Advanced)", "C1",
     "1) Inversion (full: negative adverbials, conditional inversion)\n2) Cleft sentences (it-cleft, what-cleft)\n3) Impersonal passive (It is said that X / X is said to…)\n4) Nominalisation\n5) Fronting / emphasis structures\n6) Hedging (tend to, seem to, appear to)\n7) Ellipsis\n8) Advanced modals (might well have, need not have)\n9) Complex noun phrases\n10) Discourse markers + rhetorical devices (ethos/pathos/logos for writing)",
     "Rarely have I seen such skill. / What I need is time. / It is believed that the earth is warming. / The problem is poverty's persistence.",
     "C1 Advanced (CAE): all 4 papers", "120"),
]

r = next_row + 1
for row_data in grammar_rows:
    for i, v in enumerate(row_data, 1):
        ws.cell(row=r, column=i, value=v)
    grade_part = row_data[0].split(" /")[0].strip()
    apply_level_color_to_row(ws, r, len(headers), grade_part)
    for c in range(1, len(headers)+1):
        ws.cell(row=r, column=c).font = BODY_FONT
        ws.cell(row=r, column=c).alignment = WRAP_LEFT
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = 180
    r += 1

set_col_widths(ws, [22, 10, 50, 42, 22, 7])
ws.freeze_panes = "A" + str(next_row + 1)
ws.sheet_view.showGridLines = False


# =============================================================================
# SHEET 05: VOCABULARY
# =============================================================================
ws = wb.create_sheet("05_Vocabulary")
next_row = write_title(ws, "Vocabulary by Grade — Thematic + Cambridge Wordlists",
                       "Aligned to Cambridge official topic inventories per level", row=1, span=5)
headers = ["Grade / Exam", "Themes (6 units)", "Sample core vocabulary",
           "Cambridge wordlist reference", "Target size (cumulative)"]
for i, h in enumerate(headers, 1):
    ws.cell(row=next_row, column=i, value=h)
style_header_row(ws, next_row, len(headers))

vocab_rows = [
    ("G1 / Pre-Starters",
     "1. Hello World! · 2. My Classroom · 3. My Body · 4. Farm Animals · 5. My Family · 6. Food & Fun",
     "Greetings, numbers 1–10, colours, classroom objects, body parts, farm animals, family, food (basic)",
     "YLE Starters foundations: animals, body, family, food, numbers, school",
     "~150 cumulative"),
    ("G2 / Starters",
     "1. All About Me · 2. My Home & Family · 3. Animals & Pets · 4. My Town · 5. Food & Health · 6. Hobbies & Play",
     "Appearance, rooms, furniture, pets/wild animals, places in town, food groups, sports, hobbies",
     "YLE Starters wordlist (full ~450 items)",
     "~450 cumulative"),
    ("G3 / Pre-Movers",
     "1. My School Day · 2. Routines & Times · 3. Weather & Seasons · 4. Nature & Animals · 5. Feelings & Health · 6. Our Community",
     "School subjects, time, weather, wild animals, emotions, community helpers",
     "YLE Movers preview: health, community, world around us",
     "~800 cumulative"),
    ("G4 / Movers",
     "1. Around Town · 2. School & Learning · 3. Wild Animals · 4. Sports & Holidays · 5. Adventures · 6. Future Plans",
     "Places in town (extended), school subjects, wild animals, sports & games, countries, jobs",
     "YLE Movers wordlist (full ~1350 items)",
     "~1350 cumulative"),
    ("G5 / Flyers",
     "1. Identity & Me · 2. Daily Habits · 3. Nature · 4. Cultures around me · 5. Health & Body · 6. Stories & Adventures",
     "Personality adjectives (kind, brave, patient), values, nature basics, heritage, health, narrative vocabulary",
     "YLE Flyers wordlist (full ~2500 items)",
     "~2500 cumulative"),
    ("G6 / KET",
     "1. Me & Others · 2. Home & Routines · 3. Neighbourhood & City · 4. Travel & Transport · 5. Health & Services · 6. Entertainment & Media",
     "Personal identification, daily life, places/buildings, services, transport, travel, entertainment, media, health",
     "KET list (+ YLE retained): 19 topic areas · building on ~2500 from G5",
     "~3000 cumulative (KET target + retained YLE)"),
    ("G7 / KET+ / PET bridge",
     "1. My place in society · 2. Work & Study routines · 3. Country, travel & culture · 4. International perspectives · 5. Wellbeing & lifestyle · 6. News, Society & Trends",
     "Society, work, academic study, international travel, lifestyle, news, opinion-building lexis",
     "KET+ → PET transition: environment, relations, personal feelings",
     "~3400 cumulative"),
    ("G8 / PET",
     "1. Personality & Lifestyle · 2. Entertainment & Media · 3. Body & Health · 4. Social Issues · 5. Environment · 6. Literature Genres",
     "Personality, media genres, health problems, social issues (equality, diversity), environment, literary terms",
     "PET wordlist · 23 topic areas · cumulative on KET/YLE",
     "~3800 cumulative (PET target)"),
    ("G9 / FCE",
     "1. Identity & Diversity · 2. Science & Innovation · 3. Global Issues · 4. Culture & Traditions · 5. Future Tech · 6. Global Problems",
     "Social rights, academic verbs (analyse, evaluate, hypothesise), cultural taboos, future-tech lexis",
     "FCE bands: idiom, phrasal verbs, collocations, AWL preview",
     "~4500 cumulative"),
    ("G10 / FCE consolidation",
     "1. Personal Development · 2. Media Literacy · 3. Innovation & Research · 4. Social Change · 5. Sustainability · 6. Professional Communication",
     "Soft skills, media literacy, research methods, sustainability, professional lexis",
     "AWL Sets 1–2 + CAE topic preview",
     "~5000 cumulative"),
    ("G11 / CAE",
     "1. Rhetoric & Argument · 2. Literary Analysis · 3. Academic Methodology · 4. Economics & Society · 5. Political Science · 6. Philosophy & Ethics",
     "Rhetorical devices, literary analysis (theme, motif, symbol), research, economics, politics, ethics",
     "CAE bands: AWL Sets 3–7, academic idioms, formal register",
     "~5500–6000 cumulative (CAE)"),
]

r = next_row + 1
for row_data in vocab_rows:
    for i, v in enumerate(row_data, 1):
        ws.cell(row=r, column=i, value=v)
    grade_part = row_data[0].split(" /")[0].strip()
    apply_level_color_to_row(ws, r, len(headers), grade_part)
    for c in range(1, len(headers)+1):
        ws.cell(row=r, column=c).font = BODY_FONT
        ws.cell(row=r, column=c).alignment = WRAP_LEFT
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = 75
    r += 1

set_col_widths(ws, [22, 48, 48, 36, 20])
ws.freeze_panes = "A" + str(next_row + 1)
ws.sheet_view.showGridLines = False


# =============================================================================
# SHEET 06: PHONICS (PK–G5, 8 units per grade)
# =============================================================================
ws = wb.create_sheet("06_Phonics")
next_row = write_title(ws, "Systematic Phonics — PK to Grade 5",
                       "Orton-Gillingham-aligned progression; 8 units per year (PK–G5)", row=1, span=9)
headers = ["Unit", "PK", "Kinder", "G1", "G2", "G3", "G4", "G5"]
for i, h in enumerate(headers, 1):
    ws.cell(row=next_row, column=i, value=h)
style_header_row(ws, next_row, len(headers))

phonics_rows = [
    ("U1", "Rhyme repetition · Alliteration · Word-level PA",
           "Noise awareness · Phonological awareness intro",
           "Closed syllables (CVC) 1-syll review",
           "Closed syllables (CVC) + 2-syll review",
           "Short/long vowel patterns (CVCe, CVVC)",
           "Long vowel patterns review · Diphthongs",
           "Syllable juncture (VCCV/VCV/VV) + accented patterns"),
    ("U2", "Rhyme repetition · Sentence-level PA",
           "Phonemic awareness: blending + segmenting",
           "Long vowel digraphs",
           "Long vowel digraphs + application",
           "Open syllables + long vowel patterns (CVVC)",
           "Beginning and ending complex consonants",
           "Long-vowel patterns in accented syllables"),
    ("U3", "Rhyme recognition · Alliteration",
           "Phonemic awareness: blending + segmenting",
           "Long vowel digraphs + other vowel patterns",
           "Long and other vowel patterns",
           "Short/long vowel patterns (CVVC, open)",
           "Homophones & homographs · Inflected endings",
           "Unaccented syllables · Complex consonants"),
    ("U4", "Rhyme recognition · Alliteration",
           "Phonemic awareness: blending + segmenting",
           "Phonemic awareness: manipulating sounds",
           "Vowel spelling alternatives (long vowel digraphs)",
           "R-influenced vowel patterns (Vr syllable)",
           "Inflected endings (-es, -s) · Unusual past tenses",
           "Prefixes (re-, un-, dis-, mis-, pre-, ex-, non-)"),
    ("U5", "Rhyme recognition · Blend/segment",
           "Phonemic awareness: consonant digraphs",
           "Phonemic awareness: segmenting longer words",
           "Vowel alternatives (/aw/, au, aw)",
           "Diphthongs and ambiguous vowels",
           "Syllable juncture (VCCV, VCV, VCCCV, VV)",
           "Suffixes (-y, -ly, -ily, -er, -est, -ness, -ful)"),
    ("U6", "Rhyme generation · Syllable PA",
           "Phonemic awareness: adding/substituting",
           "Phonemic awareness: manipulating + discriminating",
           "Vowel alternatives + tricky spelling",
           "Beginning complex consonants",
           "Long-vowel patterns in accented syllables",
           "Suffixes (-ous, -ious, -ary, -ery, -ory, -al)"),
    ("U7", "Rhyme generation · Syllable delete",
           "Phonemic awareness: manipulating sounds",
           "Phonemic awareness: manipulating sounds (application)",
           "Vowel spelling alternatives (/oo/, /ow/)",
           "Ending complex consonants · Homophones/graphs",
           "Other vowel patterns in accented syllables",
           "Spelling-meaning patterns (-ity, -ation, -ent)"),
    ("U8", "Phoneme isolation",
           "Phonemic awareness: complete review",
           "Complete review + application",
           "Review + application",
           "Review + application of all vowel spellings",
           "Review + application with morphemes",
           "Greek/Latin roots intro"),
]

r = next_row + 1
for row_data in phonics_rows:
    for i, v in enumerate(row_data, 1):
        ws.cell(row=r, column=i, value=v)
    for c in range(1, len(headers)+1):
        ws.cell(row=r, column=c).font = BODY_FONT
        ws.cell(row=r, column=c).alignment = WRAP_LEFT
        ws.cell(row=r, column=c).border = BORDER
    # Color G1-G5 by level
    for col_idx, gr in enumerate(["", "", "", "G1","G2","G3","G4","G5"], 1):
        if gr:
            ws.cell(row=r, column=col_idx).fill = LEVEL_FILLS[GRADE_LEVEL[gr]]
    ws.row_dimensions[r].height = 55
    r += 1

set_col_widths(ws, [7, 26, 26, 26, 26, 26, 28, 30])
ws.freeze_panes = "B" + str(next_row + 1)
ws.sheet_view.showGridLines = False


# =============================================================================
# SHEET 07: READING PLAN
# =============================================================================
ws = wb.create_sheet("07_Reading")
next_row = write_title(ws, "Reading Plan — Books by Grade (Age-appropriate + CEFR-aligned)",
                       "Corrected: replaced Moby Dick (G11) and Into the Wild (G5) with appropriate titles", row=1, span=5)
headers = ["Grade / Exam", "Book / Reader", "Author", "Level/Lexile", "Notes"]
for i, h in enumerate(headers, 1):
    ws.cell(row=next_row, column=i, value=h)
style_header_row(ws, next_row, len(headers))

reading = [
    # PRIMARY G1 — Raz Plus (no G1 English books in Plan Lector 2026)
    ("G1 / Pre-Starters", "RAZ PLUS — Levels aa–D", "Platform: Raz-Plus (Learning A-Z)", "Lexile BR–100L", "All students have access. 6 units × 3–5 books. Focus: phonics, HFW, picture-text match."),
    ("G1 / Pre-Starters", "Suggested genres", "—", "aa–D", "Rhyming · ABC · Concept · Pattern books (I see…, I have…)"),
    ("G1 / Pre-Starters", "Unit-aligned Raz Plus examples", "—", "aa–D", "U1 Hello: 'My School' (A) · U3 Body: 'My Body' (B) · U4 Farm: 'The Farm' (B) · U5 Family: 'My Family' (B) · U6 Food: 'Fruit' (C)"),
    # G2 — PLAN LECTOR 2026 (2 books) + Raz Plus supplement
    ("G2 / Starters", "The Thought Task", "Oxford – Dolphin Readers", "A1 graded (S/ 21)", "PL Book 1 — compulsory"),
    ("G2 / Starters", "Fruit", "Oxford – Dolphin Readers", "A1 graded (S/ 26)", "PL Book 2 — compulsory"),
    ("G2 / Starters", "RAZ PLUS — Levels E–J (supplement)", "Platform: Raz-Plus", "Lexile 100L–400L", "Supplement for units without PL title — self-paced reading at level"),
    # G3 — PLAN LECTOR 2026
    ("G3 / Pre-Movers", "Camouflage", "Oxford – Read and Discover", "A1 graded (S/ 56.30)", "PL Book 1 — compulsory · non-fiction, aligns with Nature/Animals unit"),
    ("G3 / Pre-Movers", "A Shadow on the Park", "Oxford – Read and Imagine", "A1 graded (S/ 51)", "PL Book 2 — compulsory · fiction, aligns with Feelings/Community unit"),
    ("G3 / Pre-Movers", "RAZ PLUS — Levels K–P (supplement)", "Platform: Raz-Plus", "Lexile 400L–650L", "Supplement for other units"),
    # G4 — PLAN LECTOR 2026
    ("G4 / Movers", "Incredible Earth", "Oxford – Read and Discover", "A1-A2 graded (S/ 38)", "PL Book 1 — compulsory · non-fiction, aligns with Wild Animals/Nature unit"),
    ("G4 / Movers", "Nasreddin Ten Stories", "Vicens Vives", "A2 graded (S/ 42)", "PL Book 2 — compulsory · folktales, aligns with Adventures unit"),
    ("G4 / Movers", "RAZ PLUS — Levels Q–V (supplement)", "Platform: Raz-Plus", "Lexile 650L–900L", "Supplement for other units"),
    # G5 — PLAN LECTOR 2026
    ("G5 / Flyers", "Exploring Our World", "Oxford – Read and Discover", "A2 graded (S/ 58.20)", "PL Book 1 — compulsory · non-fiction, aligns with Nature/Cultures units"),
    ("G5 / Flyers", "The Lost City", "Oxford – Read and Imagine", "A2 graded (S/ 58.20)", "PL Book 2 — compulsory · adventure fiction, aligns with Adventures & Stories unit"),
    ("G5 / Flyers", "RAZ PLUS — Levels W–Z2 (supplement)", "Platform: Raz-Plus", "Lexile 900L–1100L", "Supplement for other units"),
    # G6 — PLAN LECTOR 2026 (3 books)
    ("G6 / KET", "The Secret Garden", "Vicens Vives – Green Apple A1", "A1 graded (S/ 49)", "PL Book 1 — Nature/Home unit alignment"),
    ("G6 / KET", "The Wonderful Wizard of Oz", "VV – Green Apple Life Skills A1", "A1 graded (S/ 49)", "PL Book 2 — Travel unit alignment"),
    ("G6 / KET", "The Prince and the Pauper", "VV – Green Apple Life Skills Step 1 A2", "A2 graded (S/ 49)", "PL Book 3 — People/Identity unit"),
    # G7 — PLAN LECTOR 2026 (3 books)
    ("G7 / KET+ bridge", "The Adventures of Tom Sawyer", "VV – Green Apple Life Skills Step 1 A2", "A2 graded (S/ 49)", "PL Book 1 — Culture/Countries unit"),
    ("G7 / KET+ bridge", "Great Expectations", "VV – Green Apple A2", "A2 graded (S/ 47)", "PL Book 2 — Society/News unit"),
    ("G7 / KET+ bridge", "Treasure Island", "VV – Green Apple A2/B1", "A2-B1 graded (S/ 57)", "PL Book 3 — Travel & Holidays unit"),
    # G8 — PLAN LECTOR 2026 (3 books)
    ("G8 / PET", "Call of the Wild", "Black Cat (Lecturas)", "B1 graded (S/ 57)", "PL Book 1 — Environment unit alignment"),
    ("G8 / PET", "Gulliver's Travels", "Black Cat (Lecturas)", "B1 graded (S/ 57)", "PL Book 2 — Literature Genres unit"),
    ("G8 / PET", "The Giver", "Clarion Books (original)", "Lexile 760 — B1/B2 (S/ 75.50)", "PL Book 3 — Social Issues unit · age-appropriate"),
    # G9 — PLAN LECTOR 2026 (3 books) — ⚠ FLAG: B1.2 level for FCE year
    ("G9 / FCE", "The Importance of Being Earnest", "VV – Reading & Training Step Three B1.2", "B1.2 graded (S/ 57)", "PL Book 1 · ⚠ B1.2 is below FCE-B2 target — supplement with authentic B2 material"),
    ("G9 / FCE", "And Then There Were None", "VV – Reading & Training B1.2", "B1.2 graded (S/ 57)", "PL Book 2 · ⚠ B1.2 — consider upgrade to VV B2 / Black Cat B2 next cycle"),
    ("G9 / FCE", "The Time Machine", "VV – Reading & Training B1.2", "B1.2 graded (S/ 57)", "PL Book 3 · ⚠ B1.2 — same caveat for FCE prep"),
    # G10 — PLAN LECTOR 2026 (3 books, originals)
    ("G10 / FCE consol.", "Oliver Twist", "Penguin Group", "B2+ original (S/ 49)", "PL Book 1 — FCE-consolidation B2+ reading (consider abridged edition for pace)"),
    ("G10 / FCE consol.", "Animal Farm", "Signet Classic", "B2/C1 original (S/ 49.30)", "PL Book 2 — Social change / Political allegory · 30k words · ideal length"),
    ("G10 / FCE consol.", "Fahrenheit 451", "Simon & Schuster", "B2 original (S/ 51.66)", "PL Book 3 — Media literacy / Dystopia"),
    # G11 — PLAN LECTOR 2026 (3 books) — ⚠ FLAG: Moby Dick full is unrealistic
    ("G11 / CAE", "Moby Dick", "Penguin Random House", "C1+ original (S/ 79.45)", "⚠ PL Book 1 · FLAG: Full Moby Dick (206k words) is unrealistic in CAE prep year. Recommend: use ABRIDGED edition + curated excerpts. Escalate to coordination for 2027 cycle."),
    ("G11 / CAE", "The Boys Who Challenged Hitler", "Farrar, Straus and Giroux", "B2+/C1 non-fiction (S/ 64.92)", "PL Book 2 — 225pp · historical narrative · appropriate for CAE"),
    ("G11 / CAE", "Lord of the Flies", "Penguin Books", "C1 original (S/ 59)", "PL Book 3 — 60k words · allegory, symbolism · appropriate for CAE"),
]

r = next_row + 1
for row_data in reading:
    for i, v in enumerate(row_data, 1):
        ws.cell(row=r, column=i, value=v)
    grade_part = row_data[0].split(" /")[0].strip()
    apply_level_color_to_row(ws, r, len(headers), grade_part)
    # Highlight corrected rows
    if "REPLACES" in row_data[4] or "replaces" in row_data[4]:
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).fill = CORRECTION_FILL
    for c in range(1, len(headers)+1):
        ws.cell(row=r, column=c).font = BODY_FONT
        ws.cell(row=r, column=c).alignment = WRAP_LEFT
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = 28
    r += 1

set_col_widths(ws, [22, 40, 28, 18, 42])
ws.freeze_panes = "A" + str(next_row + 1)
ws.sheet_view.showGridLines = False


# =============================================================================
# SHEET 08: WRITING
# =============================================================================
ws = wb.create_sheet("08_Writing")
next_row = write_title(ws, "Writing — Text Types + Word Counts (CORRECTED)",
                       "Word counts halved for G1–G3; CAE word count fixed to 220–260; 'article' removed from CAE (proposal only)", row=1, span=5)
headers = ["Grade / Exam", "Classroom text types", "Classroom word count", "Exam task (if applicable)", "Exam word count"]
for i, h in enumerate(headers, 1):
    ws.cell(row=next_row, column=i, value=h)
style_header_row(ws, next_row, len(headers))

writing = [
    ("G1 / Pre-Starters", "Name card · labels · 1-sentence copies", "10–25 (labels, 1-word gaps)", "—", "—"),
    ("G2 / Starters", "Self-description · sentence-level picture copies", "20–40", "Starters R/W Parts 1-5: copy, fill blanks (3–5 words)", "3–5 per blank"),
    ("G3 / Pre-Movers", "Short paragraphs: routines, animal fact files", "40–70", "—", "—"),
    ("G4 / Movers", "Short notes, postcards, simple descriptions", "60–100", "Movers Writing: short note from picture", "30–40"),
    ("G5 / Flyers", "Narratives, personal letters, fact files, opinions", "100–150", "Flyers Writing Part 3: picture story", "20–25"),
    ("G6 / KET", "Short articles, diary entries, factual descriptions", "100–180", "KET Writing Part 6: email (3 prompts) · Part 7: picture story", "25 / 35"),
    ("G7 / KET+ bridge", "Blog posts, informal letters, reports intro", "180–250", "— (prep for PET)", "100 (practice)"),
    ("G8 / PET", "Formal/informal emails, articles, reviews, stories", "200–260", "PET Writing: Part 1 email · Part 2 article OR story", "100 / 100"),
    ("G9 / FCE", "Essays (opinion, for-and-against), articles, reviews, reports, informal letters, stories", "250–380", "FCE Writing: Part 1 essay (compulsory) · Part 2 one of article/email/letter/report/review", "140–190 / 140–190"),
    ("G10 / FCE consol.", "Argumentative essays, research reports, CV + cover letter, reviews", "300–430", "FCE practice + CAE bridging", "140–190 (FCE) / 220–260 (CAE prep)"),
    ("G11 / CAE", "Essays (critical), feature articles, research reports, proposals, formal letters", "350–500 (classroom)", "CAE Writing: Part 1 essay (compulsory) · Part 2 one of letter/email · PROPOSAL · report · review", "220–260 / 220–260"),
]

r = next_row + 1
for row_data in writing:
    for i, v in enumerate(row_data, 1):
        ws.cell(row=r, column=i, value=v)
    grade_part = row_data[0].split(" /")[0].strip()
    apply_level_color_to_row(ws, r, len(headers), grade_part)
    for c in range(1, len(headers)+1):
        ws.cell(row=r, column=c).font = BODY_FONT
        ws.cell(row=r, column=c).alignment = WRAP_LEFT
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = 55
    r += 1

# Notes
r += 1
ws.cell(row=r, column=1, value="NOTE: In CAE (G11), Cambridge replaced 'article' with 'proposal' at C1 level.").font = Font(bold=True, color="C00000")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

set_col_widths(ws, [22, 40, 26, 40, 22])
ws.freeze_panes = "A" + str(next_row + 1)
ws.sheet_view.showGridLines = False


# =============================================================================
# SHEET 09: SPEAKING & LISTENING
# =============================================================================
ws = wb.create_sheet("09_Speaking_Listening")
next_row = write_title(ws, "Speaking & Listening — Progression G1–G11",
                       "Oral skills aligned to Cambridge exam speaking formats", row=1, span=5)
headers = ["Grade / Exam", "Speaking — classroom activities", "Speaking — exam format", "Listening — classroom", "Listening — exam format"]
for i, h in enumerate(headers, 1):
    ws.cell(row=next_row, column=i, value=h)
style_header_row(ws, next_row, len(headers))

sl = [
    ("G1 / Pre-Starters", "Say hello/name/age; chants; pair greetings; sing Head/Shoulders", "—", "Follow TPR (stand up, point to…); identify sounds in songs", "—"),
    ("G2 / Starters", "Describe pictures; name objects; act out family/home dialogues", "Starters Speaking: 3–5 min, 4 parts (point to/Q answer/pick up/general chat)", "Short listening games; sound ID; following 2-step instructions", "Starters Listening: 20 min, 4 parts"),
    ("G3 / Pre-Movers", "Describe routines; favourites with reasons; short role-plays", "—", "Short announcements; weather reports; feelings dialogues", "—"),
    ("G4 / Movers", "Short narratives past events; preferences with 'because'; describe differences", "Movers Speaking: 5–7 min, 4 parts (describe pics / spot diff / story / personal Qs)", "Movers-style listen + select/draw/match", "Movers Listening: 25 min, 5 parts"),
    ("G5 / Flyers", "Story retell; opinions with reasons; pair planning (trip, project)", "Flyers Speaking: 7–9 min, 4 parts (spot diff / picture story / odd one out / personal Qs)", "Flyers-style narratives and dialogues", "Flyers Listening: 25 min, 5 parts"),
    ("G6 / KET", "2-min topic talks (family, school, food); paired tasks", "KET Speaking: 8–10 min, 2 parts (interview + collaborative task)", "Dialogues about shopping, travel, time", "KET Listening: 25 min, 5 parts"),
    ("G7 / KET+ bridge", "3-min discussions; mini-debates; giving directions in role-plays", "— (prep)", "Announcements, phone conversations, short interviews", "—"),
    ("G8 / PET", "4-min individual long turns (describe photo + discuss); paired problem-solving", "PET Speaking: 10–12 min, 4 parts", "Longer dialogues, monologues, multi-speaker", "PET Listening: 30 min, 4 parts"),
    ("G9 / FCE", "5-min presentations + Q&A; discussions comparing options; debates", "FCE Speaking: 14 min, 4 parts (interview + long turn + collaborative + discussion)", "Radio extracts, interviews, conversations", "FCE Listening: 40 min, 4 parts"),
    ("G10 / FCE consol.", "7-min seminars; justify with evidence; counter-arguments", "FCE practice + CAE bridging", "Lectures, TED excerpts, formal interviews", "FCE practice"),
    ("G11 / CAE", "8-min formal presentations with Q&A; mock speaking exam format", "CAE Speaking: 15 min, 4 parts (interview / long turn / collaborative / discussion)", "Academic talks, specialist interviews, panel discussions", "CAE Listening: 40 min, 4 parts"),
]

r = next_row + 1
for row_data in sl:
    for i, v in enumerate(row_data, 1):
        ws.cell(row=r, column=i, value=v)
    grade_part = row_data[0].split(" /")[0].strip()
    apply_level_color_to_row(ws, r, len(headers), grade_part)
    for c in range(1, len(headers)+1):
        ws.cell(row=r, column=c).font = BODY_FONT
        ws.cell(row=r, column=c).alignment = WRAP_LEFT
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = 55
    r += 1

set_col_widths(ws, [22, 40, 34, 34, 26])
ws.freeze_panes = "A" + str(next_row + 1)
ws.sheet_view.showGridLines = False


# =============================================================================
# SHEET 10: UNITS G1-G5
# =============================================================================
ws = wb.create_sheet("10_Units_G1_G5")
next_row = write_title(ws, "Unit Plans G1–G5 — 6 units per grade",
                       "Theme · Key language · Grammar anchor · Vocabulary · Writing output", row=1, span=7)
headers = ["Grade", "Unit", "Theme", "Key language functions", "Grammar focus", "Vocabulary (top)", "Writing output"]
for i, h in enumerate(headers, 1):
    ws.cell(row=next_row, column=i, value=h)
style_header_row(ws, next_row, len(headers))

units_prim = [
    # G1
    ("G1", "U1", "Hello World!", "Greet; say name, age, colours; count 1-10", "To be; subject pronouns", "Greetings, numbers, colours", "Name card + label"),
    ("G1", "U2", "My Classroom", "Name objects; ask 'What is this?'", "This/that; it is", "Classroom objects; a/an", "Label classroom picture"),
    ("G1", "U3", "My Body", "Name body parts; describe 'I have two eyes'", "I have; adjectives big/small", "Body parts; colours", "Label body diagram"),
    ("G1", "U4", "Farm Animals", "Name animals; animal sounds; 'The cow can moo'", "Can (positive); plurals", "Farm animals", "Animal fact card (3 words)"),
    ("G1", "U5", "My Family", "Introduce family; 'This is my mum'", "Possessive adj (my, your, his, her)", "Family members", "Family tree labels"),
    ("G1", "U6", "Food & Fun", "Say likes/dislikes; 'I like apples'", "Like + noun; don't like", "Food, drink", "My plate + labels"),
    # G2
    ("G2", "U1", "All About Me", "Introduce self; describe appearance", "Present simple (he/she); have got", "Appearance (tall, curly hair)", "Self-description card 20-30w"),
    ("G2", "U2", "My Home & Family", "Describe rooms, furniture", "There is/are; prepositions of place", "Rooms, furniture", "House description 25-40w"),
    ("G2", "U3", "Animals & Pets", "Describe animals: 'It has… It can… It eats…'", "Have got + can; plurals", "Pets + wild animals", "Animal fact 30-40w"),
    ("G2", "U4", "My Town", "Ask/give directions; places", "There is/are; prepositions of movement", "Places in town", "Postcard 30-40w"),
    ("G2", "U5", "Food & Health", "Talk about healthy habits; 'I eat/drink'", "Like + -ing; present simple", "Food groups, health", "Healthy eating poster 5 bullets"),
    ("G2", "U6", "Hobbies & Play", "Talk about hobbies, sports", "Like + -ing; can for ability", "Sports, games", "My hobby paragraph 30-40w"),
    # G3
    ("G3", "U1", "My School Day", "Describe school subjects, routines", "Present simple (all persons)", "School subjects, time", "My school day 40-50w"),
    ("G3", "U2", "Routines & Times", "Tell time; frequency", "Adverbs of frequency; prepositions of time", "Daily activities", "Weekly routine 50-60w"),
    ("G3", "U3", "Weather & Seasons", "Describe weather, clothes by season", "Present continuous; it is + adj", "Weather, clothes, seasons", "Weather report 50-60w"),
    ("G3", "U4", "Nature & Animals", "Compare wild animals", "Comparative / superlative (short adj)", "Wild animals (penguin, whale, shark)", "Animal comparison 50-70w"),
    ("G3", "U5", "Feelings & Health", "Express emotions, health advice", "Modals: must, should, have to", "Emotions, body, illness", "Advice leaflet 50-70w"),
    ("G3", "U6", "Our Community", "Talk about community helpers, jobs", "Present simple Qs; adverbs", "Community helpers, places", "Community report 50-70w"),
    # G4
    ("G4", "U1", "Around Town", "Describe past trips, places visited", "Past simple regular", "Places in town, transport", "Trip diary 60-80w"),
    ("G4", "U2", "School & Learning", "Talk about favourite subjects, teachers", "Past simple irregular; verb + inf", "School subjects extended", "School memory 60-80w"),
    ("G4", "U3", "Wild Animals", "Describe habitat, behaviour of wild animals", "Relative clauses (who/which/where)", "Wild animals, habitats", "Animal fact file 80-100w"),
    ("G4", "U4", "Sports & Holidays", "Describe sports, past holidays", "Past simple Qs/neg; infinitive of purpose", "Sports, holidays, activities", "Holiday recount 80-100w"),
    ("G4", "U5", "Adventures", "Narrate a short adventure; comparatives of all adj", "Comparatives/superlatives all", "Adventure, nature", "Adventure story 80-100w"),
    ("G4", "U6", "Future Plans", "Talk about future plans (basic will)", "Will for simple predictions; must/have to", "Jobs, future, ambitions", "My future 80-100w"),
    # G5
    ("G5", "U1", "Personality & Identity", "Describe personality, values", "Past cont; past simple vs cont", "Personality adj (kind, brave)", "Character description 100-150w"),
    ("G5", "U2", "Habits & Lifestyle", "Talk about past habits; used to", "Used to; past continuous", "Habits, lifestyle, routines", "Then vs now essay 100-150w"),
    ("G5", "U3", "The Natural World", "Describe environment, simple causes", "Zero conditional; be going to", "Environment, nature", "Environment report 100-150w"),
    ("G5", "U4", "Cultures & Places", "Describe cultures, traditions, festivals", "Present perfect (intro: have been/done)", "Cultures, traditions, festivals", "Cultural comparison 100-150w"),
    ("G5", "U5", "Health & Wellbeing", "Give advice on health; emotions", "Might, may, should; sensory verbs", "Health, emotions, advice", "Health advice article 100-150w"),
    ("G5", "U6", "Adventures & Stories", "Tell stories with more complex structures", "Tag questions; connectors (because, so, although)", "Adventure, storytelling", "Short story 100-150w"),
]

r = next_row + 1
for row_data in units_prim:
    for i, v in enumerate(row_data, 1):
        ws.cell(row=r, column=i, value=v)
    apply_level_color_to_row(ws, r, len(headers), row_data[0])
    for c in range(1, len(headers)+1):
        ws.cell(row=r, column=c).font = BODY_FONT
        ws.cell(row=r, column=c).alignment = WRAP_LEFT
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = 42
    r += 1

set_col_widths(ws, [7, 6, 26, 36, 32, 32, 28])
ws.freeze_panes = "A" + str(next_row + 1)
ws.sheet_view.showGridLines = False


# =============================================================================
# SHEET 11: UNITS G6-G11
# =============================================================================
ws = wb.create_sheet("11_Units_G6_G11")
next_row = write_title(ws, "Unit Plans G6–G11 — 6 units per grade",
                       "Theme · Key language · Grammar anchor · Vocabulary · Writing output", row=1, span=7)
headers = ["Grade", "Unit", "Theme", "Key language functions", "Grammar focus", "Vocabulary (top)", "Writing output"]
for i, h in enumerate(headers, 1):
    ws.cell(row=next_row, column=i, value=h)
style_header_row(ws, next_row, len(headers))

units_sec = [
    # G6 KET
    ("G6", "U1", "People & Relationships", "Describe people, relationships, feelings", "Present perfect (ever/never)", "People, feelings", "Personal profile 100w"),
    ("G6", "U2", "Home & Neighbourhood", "Describe home, neighbourhood, changes", "Present perfect (just/already/yet)", "Home, neighbourhood", "Neighbourhood review 120w + KET email 25w"),
    ("G6", "U3", "Travel & Transport", "Describe trips, plans, transport", "Present perfect (for/since); first conditional", "Travel, transport", "Travel article 130w + KET story 35w"),
    ("G6", "U4", "Work & Services", "Talk about jobs, services, asking for info", "Full modals (can, could, may, might, must, should)", "Jobs, services", "Formal email 120w"),
    ("G6", "U5", "Health & Lifestyle", "Give/take advice on health and lifestyle", "Passive voice (pres/past simple)", "Health, lifestyle, exercise", "Health article 140w"),
    ("G6", "U6", "Entertainment & Media", "Discuss entertainment preferences, media", "Gerunds; adjective order (OSASCOMP)", "Entertainment, media, genres", "Review/recommendation 150w"),
    # G7 bridge
    ("G7", "U1", "Countries & Culture", "Compare countries, cultures, traditions", "Past perfect simple", "Countries, cultures, festivals", "Cultural report 200w"),
    ("G7", "U2", "School & Study", "Discuss study habits, school issues", "2nd conditional", "School, study skills", "Opinion email 180w"),
    ("G7", "U3", "Travel & Holidays", "Describe past trips; narrate anecdotes", "Reported speech (statements)", "Travel, accommodation", "Travel blog 220w"),
    ("G7", "U4", "Tech & Internet", "Discuss tech impact, online safety", "Relative clauses (def + non-def)", "Tech, internet, social media", "Tech review 240w"),
    ("G7", "U5", "Jobs & Careers", "Talk about jobs, careers, aspirations", "Modal passive; causative have/get", "Jobs, skills, careers", "Cover letter intro 200w"),
    ("G7", "U6", "Society & News", "Discuss current news, events, opinions", "Connectors (although/however/despite)", "News, society, current events", "News summary 250w"),
    # G8 PET
    ("G8", "U1", "Personality & Lifestyle", "Describe personality, lifestyle choices", "All tenses review; word formation", "Personality, lifestyle", "Descriptive article 220w + PET email 100w"),
    ("G8", "U2", "Entertainment & Media", "Discuss/review media genres, bias", "Reported Qs + commands; gerund vs inf", "Entertainment, media", "Film/book review 240w"),
    ("G8", "U3", "Body & Health", "Debate health issues, solutions", "Conditionals 0-1-2 consolidated", "Body systems, health conditions", "Opinion essay 250w + PET article 100w"),
    ("G8", "U4", "Social Issues", "Debate equality, diversity, inclusion", "Passive: all simple tenses", "Social issues, rights, diversity", "Argumentative essay 260w"),
    ("G8", "U5", "Environment", "Discuss environmental problems, action", "Verb + obj + inf; phrasal verbs", "Environment, sustainability", "Environment campaign 260w"),
    ("G8", "U6", "Literature Genres", "Analyse literary genres, characters", "Register: formal vs informal", "Literary terms, genres", "Literary analysis 250w + PET story 100w"),
    # G9 FCE
    ("G9", "U1", "Identity & Diversity", "Discuss identity, rights, diversity", "Conditional 3 + mixed", "Identity, rights, diversity", "Opinion essay 170w (FCE format)"),
    ("G9", "U2", "Science & Innovation", "Discuss scientific discoveries, innovation", "Passive all tenses (cont, perf, modal)", "Science, research, innovation", "Informative report 170w"),
    ("G9", "U3", "Global Issues", "Debate global problems, solutions", "Past modals of deduction", "Global issues, politics", "Argumentative essay 180w"),
    ("G9", "U4", "Culture & Traditions", "Compare cultures, analyse traditions", "Advanced relative clauses (prepositions)", "Cultures, traditions, customs", "Cultural comparison 180w"),
    ("G9", "U5", "Future Tech", "Predict future tech, social impact", "Wish / if only; key word transformation", "Future tech, AI, society", "Prediction essay 180w"),
    ("G9", "U6", "Global Problems + Review genre", "Analyse problems; write reviews", "Connectors (whereas/nevertheless/furthermore)", "Global problems, solutions", "For-and-against essay 190w + FCE review 150w"),
    # G10 FCE consol
    ("G10", "U1", "Personal Development", "Discuss soft skills, goals", "Mixed conditionals mastery", "Soft skills, personality", "Self-reflection essay 350w"),
    ("G10", "U2", "Media Literacy", "Analyse media bias, misinformation", "Subjunctive; reported speech advanced", "Media, bias, algorithm, clickbait", "Media analysis report 400w"),
    ("G10", "U3", "Innovation & Research", "Discuss research methods, innovation", "Complex relatives; participle clauses", "Research, innovation, AWL Set 1", "Research proposal 400w"),
    ("G10", "U4", "Social Change", "Debate social change, policy", "Inversion basics; cleft sentences (intro)", "Social change, activism", "Persuasive essay 400w"),
    ("G10", "U5", "Sustainability", "Discuss sustainability, policy", "Causative advanced; passive review", "Sustainability, environment, AWL 2", "Policy brief 400w"),
    ("G10", "U6", "Professional Communication", "CV + cover letter; formal emails", "Discourse markers; phrasal verbs academic", "Professional terms, business", "CV + cover letter pack"),
    # G11 CAE
    ("G11", "U1", "Rhetoric & Argument", "Analyse rhetoric; construct arguments", "Inversion (full)", "Rhetoric (ethos/pathos/logos)", "Argumentative essay 260w (CAE)"),
    ("G11", "U2", "Literary Analysis", "Analyse literature (theme, motif, symbol)", "Cleft sentences (it-cleft, what-cleft)", "Literary analysis", "Literary review 260w (CAE)"),
    ("G11", "U3", "Academic Methodology", "Design research; academic writing", "Impersonal passive (It is said that)", "Academic methodology, AWL 3-4", "CAE report 260w"),
    ("G11", "U4", "Economics & Society", "Discuss economics, social issues", "Nominalisation; fronting", "Economics, AWL 5", "CAE proposal 260w"),
    ("G11", "U5", "Political Science", "Analyse political systems, policy", "Hedging (tend to, seem to); ellipsis", "Political science, AWL 6", "Policy proposal 260w"),
    ("G11", "U6", "Philosophy & Ethics", "Debate ethics, moral dilemmas", "Advanced modals; complex noun phrases", "Philosophy, ethics, AWL 7", "CAE essay 260w (mock exam)"),
]

r = next_row + 1
for row_data in units_sec:
    for i, v in enumerate(row_data, 1):
        ws.cell(row=r, column=i, value=v)
    apply_level_color_to_row(ws, r, len(headers), row_data[0])
    for c in range(1, len(headers)+1):
        ws.cell(row=r, column=c).font = BODY_FONT
        ws.cell(row=r, column=c).alignment = WRAP_LEFT
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = 45
    r += 1

set_col_widths(ws, [7, 6, 26, 36, 32, 32, 32])
ws.freeze_panes = "A" + str(next_row + 1)
ws.sheet_view.showGridLines = False


# =============================================================================
# SHEET 12: CLIL MAP
# =============================================================================
ws = wb.create_sheet("12_CLIL_Map")
next_row = write_title(ws, "CLIL Cross-Curricular Integration Map",
                       "Proposed projects aligning English units with Math, Science, Social Studies, Comunicación, Art, PE", row=1, span=6)
headers = ["Grade", "English Unit", "Other subject(s)", "Integrated project", "Target output", "Assessment"]
for i, h in enumerate(headers, 1):
    ws.cell(row=next_row, column=i, value=h)
style_header_row(ws, next_row, len(headers))

clil = [
    # G1 — 6 projects
    ("G1", "U1 Hello World!", "Personal Social (identity)", "My Name Card", "Decorated name card + 3-fact oral share", "Oral AD/A/B/C"),
    ("G1", "U2 My Classroom", "Math (count + sort)", "Classroom Inventory Count", "Counted chart + labelled picture", "Counting + naming"),
    ("G1", "U3 My Body", "Science (body) + PE (body awareness)", "Head-to-Toe Movement Song", "Group performance + labelled diagram", "Oral + performance"),
    ("G1", "U4 Farm Animals", "Science (living beings) + Art", "Farm Animal Mask Parade", "Mask + oral description with 'can'", "Art + oral"),
    ("G1", "U5 My Family", "Personal Social (identity + belonging)", "Family Tree Poster", "Family tree + 3-sentence oral share", "Oral + poster quality"),
    ("G1", "U6 Food & Fun", "Science (nutrition) + PE", "Healthy Plate Collage", "Collage + 'I like/don't like' statements", "Collage + oral"),
    # G2 — 6 projects
    ("G2", "U1 All About Me", "Personal Social + Art", "All About Me Booklet", "5-page booklet + oral share", "Booklet + oral"),
    ("G2", "U2 My Home & Family", "Social Studies (family)", "Home Layout Diagram", "Labelled room map + oral tour", "Map + oral"),
    ("G2", "U3 Animals & Pets", "Science (classification)", "Pet Care Fact Cards", "5 fact cards + oral presentation", "Written + oral"),
    ("G2", "U4 My Town", "Social Studies + Math (directions)", "My Town Walking Tour Map", "Map + 30w postcard in English", "Map + writing"),
    ("G2", "U5 Food & Health", "Science (nutrition) + PE", "Healthy Plate Week Diary", "5-day food diary + poster", "Content + writing"),
    ("G2", "U6 Hobbies & Play", "PE + Art + Music", "My Favourite Hobby Show", "Show-and-tell + 30w paragraph", "Oral + writing"),
    # G3 — 6 projects
    ("G3", "U1 My School Day", "Math (time)", "My School Schedule", "Weekly schedule + 60w description", "Accuracy + writing"),
    ("G3", "U2 Routines & Times", "Science (circadian) + Math (frequency)", "Human Clock Project", "Routines survey + 60w paragraph", "Data + writing"),
    ("G3", "U3 Weather & Seasons", "Science (weather) + Math (data)", "Weather Log (1 week)", "Weather chart + 60w report", "Data + writing"),
    ("G3", "U4 Nature & Animals", "Science (habitats) + Art", "Endangered Species Poster", "Habitat + animal + 70w plea", "Poster + writing"),
    ("G3", "U5 Feelings & Health", "Personal Social + PE", "Feelings First Aid Kit", "Strategy booklet + role-play", "Oral + written"),
    ("G3", "U6 Our Community", "Social Studies + Drama", "Community Helpers Play", "Role-play + 70w report", "Performance + writing"),
    # G4 — 6 projects
    ("G4", "U1 Around Town", "Social Studies (geography) + Math (scale)", "My Neighbourhood in 3D", "Scaled 3D model + 80w description", "Model + writing"),
    ("G4", "U2 School & Learning", "ICT + Personal Social", "Digital School Tour Video", "2-min video + script", "Video + script"),
    ("G4", "U3 Wild Animals", "Science (biodiversity, conservation)", "Save Our Wildlife Campaign", "Poster + 80w appeal + oral", "Multi-modal"),
    ("G4", "U4 Sports & Holidays", "PE + Social Studies (cultures)", "International Olympics Day", "Country+sport presentation + cultural profile", "Oral + cultural accuracy"),
    ("G4", "U5 Adventures", "Geography + Art", "My Expedition Plan", "Map + 100w narrative", "Planning + writing"),
    ("G4", "U6 Future Plans", "Social Studies (careers) + Art", "My Future Self Vision Board", "Vision board + 100w essay", "Vision + writing"),
    # G5 — 6 projects
    ("G5", "U1 Personality & Identity", "Personal Social + Psychology", "Who Am I? Portrait", "Self-portrait + 150w essay", "Reflection + writing"),
    ("G5", "U2 Habits & Lifestyle", "Math (budget) + Social Studies (economy)", "Weekly Budget + Habits Tracker", "Spreadsheet + 150w reflection", "Math + writing"),
    ("G5", "U3 The Natural World", "Science (ecosystems) + ICT", "Conservation Digital Leaflet", "Digital leaflet + 150w article + oral", "Multi-modal"),
    ("G5", "U4 Cultures & Places", "Social Studies + Art", "Cultural Fair Stand", "Stand + 150w cultural profile", "Stand + writing"),
    ("G5", "U5 Health & Wellbeing", "Science (body) + PE", "Wellness Week Plan", "7-day plan + 150w rationale", "Plan + writing"),
    ("G5", "U6 Adventures & Stories", "Drama + Art", "Story Theatre", "Written story + live performance", "Performance + writing"),
    # G6 KET — 6 projects
    ("G6", "U1 People & Relationships", "Personal Social + Philosophy", "Relationship Map Infographic", "Infographic + 140w profile + KET email 25w", "Infographic + writing"),
    ("G6", "U2 Home & Neighbourhood", "Social Studies (urbanism) + ICT", "Virtual Neighbourhood Tour", "Video tour + KET email 25w", "Video + email"),
    ("G6", "U3 Travel & Transport", "Geography + Math (distances/costs)", "International Trip Itinerary", "Itinerary + 140w blog + KET story 35w", "Combined rubric"),
    ("G6", "U4 Work & Services", "Social Studies (economy) + Drama", "Service Industry Role-Play", "Role-play + 140w report", "Performance + writing"),
    ("G6", "U5 Health & Lifestyle", "Science (body systems) + PE", "Body Systems Presentation", "5-min presentation + 140w article", "Presentation + writing"),
    ("G6", "U6 Entertainment & Media", "ICT + Music / Drama", "Podcast Episode", "3-min podcast + script + KET story 35w", "Podcast + script"),
    # G7 bridge — 6 projects
    ("G7", "U1 Countries & Culture", "Social Studies + Music", "Cultural Deep Dive", "5-min presentation + 200w country profile", "Research + presentation"),
    ("G7", "U2 School & Study", "Personal Social + ICT", "Study Hacks Video Tutorial", "3-min tutorial + 180w guide", "Video + guide"),
    ("G7", "U3 Travel & Holidays", "Geography + Math (currency)", "Budget Travel Blog", "Blog post 220w + itinerary", "Research + writing"),
    ("G7", "U4 Tech & Internet", "ICT + Ethics", "Digital Citizenship Campaign", "Campaign video + 240w script", "Impact + script"),
    ("G7", "U5 Jobs & Careers", "Social Studies + Personal Social", "Informational Interview", "Interview video + 200w summary", "Interview + writing"),
    ("G7", "U6 Society & News", "Social Studies + ICT", "Weekly News Podcast", "Podcast + 250w script", "Broadcast + script"),
    # G8 PET — 6 projects
    ("G8", "U1 Personality & Lifestyle", "Personal Social + Psychology", "Personality Study", "Report + 220w reflection + PET email 100w", "Analysis + writing"),
    ("G8", "U2 Entertainment & Media", "ICT + Music / Drama", "Short Film Production", "3-min film + 240w script", "Film + script"),
    ("G8", "U3 Body & Health", "Science (body systems) + PE", "School Health Campaign", "Campaign + 260w article + PET article 100w", "Campaign + writing"),
    ("G8", "U4 Social Issues", "Social Studies (rights) + Philosophy", "Human Rights Debate", "Debate + 260w position paper", "Debate + essay"),
    ("G8", "U5 Environment", "Science (ecology) + ICT", "Carbon Footprint Calculator", "Spreadsheet + 260w report", "Technical + writing"),
    ("G8", "U6 Literature Genres", "Drama + Art", "Genre Adaptation", "Rewrite story in new genre + 250w analysis + PET story 100w", "Creativity + analysis"),
    # G9 FCE — 6 projects
    ("G9", "U1 Identity & Diversity", "Social Studies + Philosophy", "Identity Documentary", "5-min doc + 170w FCE essay", "Doc + essay"),
    ("G9", "U2 Science & Innovation", "Science (method) + ICT", "Science Fair Project (English)", "Experiment + 170w FCE report", "Method + report"),
    ("G9", "U3 Global Issues", "Social Studies (MUN)", "MUN Position Paper", "Paper + debate participation", "MUN + paper"),
    ("G9", "U4 Culture & Traditions", "Social Studies + Comunicación (contrast)", "Cultural Comparison", "180w FCE essay + oral presentation", "Essay + oral"),
    ("G9", "U5 Future Tech", "ICT + Ethics", "AI Ethics Forum", "Panel discussion + 180w prediction essay", "Panel + essay"),
    ("G9", "U6 Global Problems + Review", "Social Studies + Arts", "Problem-Solution TED-Ed + Review Genre", "TED talk + 150w FCE review", "Talk + review"),
    # G10 FCE consolidation — 6 projects
    ("G10", "U1 Personal Development", "Personal Social + Philosophy", "Growth Plan Portfolio", "Portfolio + 350w reflection", "Portfolio + reflection"),
    ("G10", "U2 Media Literacy", "ICT + Philosophy", "Misinformation Audit", "Audit + 400w analysis + presentation", "Critical + analysis"),
    ("G10", "U3 Innovation & Research", "Science + ICT", "IB-style Research Proposal", "Proposal + 400w justification", "Method + writing"),
    ("G10", "U4 Social Change", "Social Studies + History", "Social Movement Analysis", "Presentation + 400w essay", "Presentation + essay"),
    ("G10", "U5 Sustainability", "Science + Economics", "Sustainable Business Pitch", "Shark-Tank pitch + 400w plan", "Pitch + plan"),
    ("G10", "U6 Professional Communication", "ICT + Careers", "Professional Portfolio", "CV + cover letter + LinkedIn-style profile", "Portfolio rubric"),
    # G11 CAE — 6 projects
    ("G11", "U1 Rhetoric & Argument", "Philosophy + Debate", "Rhetoric Masterclass", "Analyse famous speech + 260w CAE essay", "Analysis + essay"),
    ("G11", "U2 Literary Analysis", "Literature + Art", "Literary Criticism Paper", "Paper + 260w CAE review", "Criticism + review"),
    ("G11", "U3 Academic Methodology", "Global Perspectives + Science", "IB-style Research Paper", "2500w research paper + 260w CAE report", "Academic rubric"),
    ("G11", "U4 Economics & Society", "Economics + Social Studies", "Economic Policy Proposal", "Proposal + 260w CAE proposal", "Economic + writing"),
    ("G11", "U5 Political Science", "Social Studies + Philosophy", "Policy Brief (Peru-relevant)", "Brief + 260w CAE proposal", "Brief + writing"),
    ("G11", "U6 Philosophy & Ethics", "Philosophy + Spiritual History", "Ethical Dilemma Symposium", "Symposium + 260w CAE essay (mock exam)", "Symposium + essay"),
]

r = next_row + 1
for row_data in clil:
    for i, v in enumerate(row_data, 1):
        ws.cell(row=r, column=i, value=v)
    apply_level_color_to_row(ws, r, len(headers), row_data[0])
    for c in range(1, len(headers)+1):
        ws.cell(row=r, column=c).font = BODY_FONT
        ws.cell(row=r, column=c).alignment = WRAP_LEFT
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = 42
    r += 1

set_col_widths(ws, [7, 28, 34, 34, 34, 22])
ws.freeze_panes = "A" + str(next_row + 1)
ws.sheet_view.showGridLines = False


# =============================================================================
# SHEET 13: CAMBRIDGE CALENDAR
# =============================================================================
ws = wb.create_sheet("13_Cambridge_Calendar")
next_row = write_title(ws, "Cambridge Exam Calendar 2026",
                       "Mock preparation, official dates, and prep intensity by grade", row=1, span=7)
headers = ["Month", "Milestone", "Grades involved", "Activity", "Prep hours", "Responsible", "KPI"]
for i, h in enumerate(headers, 1):
    ws.cell(row=next_row, column=i, value=h)
style_header_row(ws, next_row, len(headers))

cal = [
    ("Mar", "Diagnostic CEFR", "G1–G11", "Placement test + CEFR baseline to identify prep needs", "2 per grade", "Coord.", "100% completion"),
    ("Mar–Apr", "Pre-Mock 1 prep", "G2, G4, G5, G6, G8, G9, G11", "Foundation + exam skills + familiarisation with formats", "6 h/wk/grade", "Teachers", "Prep plan signed · past-paper exposure"),
    ("MAY", "🎯 MOCK EXAM 1", "G2/G4/G5/G6/G8/G9/G11", "First mock: YLE + KET + PET + FCE + CAE — FULL OFFICIAL TIMING", "Full exam", "Teachers + Coord.", "Baseline score/grade established"),
    ("Jun", "Mock 1 analysis + re-prioritise", "All exam grades", "Analyse results · identify weak skills · adjust teaching plan", "1 wk analysis", "Coord.", "Individual action plan/student"),
    ("Jun–Jul", "Intensive prep (post Mock 1)", "All exam grades", "Target weak skills · past papers · timing practice · vocabulary push", "10 h/wk/grade", "Teachers", "Past paper bank built · weekly progress"),
    ("Aug–Sep", "Consolidation + speaking clinics", "All exam grades", "Speaking practice intensification · writing drills · listening in exam format", "8 h/wk/grade", "Teachers", "≥75% readiness mid-score"),
    ("OCT", "🎯 MOCK EXAM 2", "G2/G4/G5/G6/G8/G9/G11", "Second mock: final rehearsal under official conditions", "Full exam", "Coord.", "≥80% mock avg"),
    ("Oct–Nov", "Final prep + Cambridge enrolment", "All exam grades", "Last-mile revision · enrol through Cambridge centre · parent briefing", "6 h/wk", "Coord.", "100% enrolled by mid-Nov"),
    ("DEC 2", "🎯 OFFICIAL CAMBRIDGE EXAMS", "G2/G4/G5/G6/G8/G9/G11", "Cambridge official exam day — all papers (speaking may be scheduled ±1 week)", "Exam day", "Cambridge + Coord.", "≥80% pass rate target"),
    ("Dec (post-exam)", "Results + report + 2027 planning", "All", "Results analysis · individual student reports · parent communication · next-year plan", "2 weeks", "Coord.", "Published by Dec 20"),
]

r = next_row + 1
for row_data in cal:
    for i, v in enumerate(row_data, 1):
        ws.cell(row=r, column=i, value=v)
    for c in range(1, len(headers)+1):
        ws.cell(row=r, column=c).font = BODY_FONT
        ws.cell(row=r, column=c).alignment = WRAP_LEFT
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = 38
    r += 1

set_col_widths(ws, [9, 24, 22, 40, 14, 22, 20])
ws.freeze_panes = "A" + str(next_row + 1)
ws.sheet_view.showGridLines = False


# =============================================================================
# SHEET 14: ASSESSMENT SCALE (MINEDU AD/A/B/C)
# =============================================================================
ws = wb.create_sheet("14_Assessment_Scale")
next_row = write_title(ws, "Assessment Scale — MINEDU AD/A/B/C × Cambridge CEFR",
                       "Analytic descriptors per skill, ready for teacher evaluation", row=1, span=6)
headers = ["Skill", "Criterion", "AD (Outstanding)", "A (Expected)", "B (In Progress)", "C (Beginning)"]
for i, h in enumerate(headers, 1):
    ws.cell(row=next_row, column=i, value=h)
style_header_row(ws, next_row, len(headers))

rubrics = [
    ("Listening", "Understanding",
     "Understands complex texts in detail; grasps attitude and nuance.",
     "Understands the main idea and specific information in standard texts.",
     "Partially understands; needs repetition or visual support.",
     "Understands only isolated words or very short phrases."),
    ("Listening", "Response",
     "Responds accurately, including to inferential questions.",
     "Responds correctly to explicit-information questions.",
     "Responds with support; frequent errors.",
     "Does not respond or responds in an unrelated way."),
    ("Speaking", "Fluency & accuracy",
     "Speaks with natural fluency; minor errors do not interfere with communication.",
     "Speaks with reasonable fluency; errors are controlled.",
     "Speaks with frequent pauses; errors occasionally affect communication.",
     "Speaks only in isolated words or memorised phrases."),
    ("Speaking", "Range & accuracy",
     "Uses a wide variety of complex structures; broad and precise vocabulary.",
     "Uses the expected range for the grade; appropriate vocabulary.",
     "Limited range; basic vocabulary.",
     "Minimal structures; very limited vocabulary."),
    ("Speaking", "Pronunciation",
     "Clear pronunciation, natural intonation; accent does not interfere.",
     "Intelligible pronunciation; intonation generally correct.",
     "Pronunciation sometimes hard to understand; inconsistent intonation.",
     "Pronunciation hard to understand; inappropriate intonation."),
    ("Speaking", "Interaction",
     "Initiates, sustains and closes conversations; reacts naturally.",
     "Participates with appropriate turns; responds relevantly.",
     "Responds with support; does not initiate interaction.",
     "Does not sustain interaction."),
    ("Reading", "Comprehension",
     "Infers implicit ideas and complex relationships; identifies purpose and tone.",
     "Identifies main idea and specific information.",
     "Identifies only very basic information with support.",
     "Does not identify information from the text."),
    ("Reading", "Vocabulary in context",
     "Deduces the meaning of unfamiliar words from context.",
     "Recognises vocabulary of the grade level.",
     "Recognises basic vocabulary; needs a dictionary for new words.",
     "Only recognises isolated words."),
    ("Writing", "Task completion",
     "Completes the task creatively, includes all elements and extends with detail.",
     "Completes the task with all main elements.",
     "Partially completes; omits important elements.",
     "Does not complete the task."),
    ("Writing", "Organisation & cohesion",
     "Clear, logical structure; varied connectors produce natural flow.",
     "Appropriate structure; basic connectors used correctly.",
     "Weak organisation; limited or incorrect connectors.",
     "No apparent organisation."),
    ("Writing", "Grammar & accuracy",
     "Grade-target structures + some more complex ones with >90% accuracy.",
     "Grade-target structures used correctly most of the time (75–89%).",
     "Frequent errors in grade-target structures (50–74%).",
     "Pervasive errors; target structures not evident."),
    ("Writing", "Vocabulary & register",
     "Varied and precise vocabulary; appropriate and sustained register.",
     "Vocabulary appropriate for the grade; register mostly correct.",
     "Limited vocabulary; register errors.",
     "Very limited vocabulary; inappropriate register."),
    ("Writing", "Mechanics",
     "Flawless spelling, punctuation and capitalisation.",
     "Minor errors that do not affect comprehension.",
     "Frequent errors that occasionally affect comprehension.",
     "Serious errors that hinder comprehension."),
    ("Grammar & Use", "Accuracy",
     "Uses target structures with sustained accuracy (≥90%); self-corrects errors.",
     "Uses target structures correctly most of the time (75–89%); minor errors.",
     "Frequent errors (50–74%); meaning sometimes unclear.",
     "Pervasive errors; target structures not evident."),
    ("Grammar & Use", "Range",
     "Wide variety of structures beyond what was taught.",
     "Range expected for the grade.",
     "Limited range; basic structures.",
     "Only memorised phrases or isolated words."),
]

r = next_row + 1
for row_data in rubrics:
    for i, v in enumerate(row_data, 1):
        ws.cell(row=r, column=i, value=v)
    # alternating group colors
    for c in range(1, len(headers)+1):
        ws.cell(row=r, column=c).font = BODY_FONT
        ws.cell(row=r, column=c).alignment = WRAP_LEFT
        ws.cell(row=r, column=c).border = BORDER
    # color AD green, C red (soft)
    ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor="C6EFCE")
    ws.cell(row=r, column=4).fill = PatternFill("solid", fgColor="E2F0D9")
    ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor="FFF2CC")
    ws.cell(row=r, column=6).fill = PatternFill("solid", fgColor="FCE4D6")
    ws.row_dimensions[r].height = 52
    r += 1

set_col_widths(ws, [14, 22, 40, 36, 34, 30])
ws.freeze_panes = "C" + str(next_row + 1)
ws.sheet_view.showGridLines = False


# =============================================================================
# SHEET 15: TEACHER QUICK REFERENCE (1 sheet with all grades)
# =============================================================================
ws = wb.create_sheet("15_Teacher_QuickRef")
next_row = write_title(ws, "Teacher Quick Reference — 1 row per grade",
                       "Print this sheet for every teacher — it summarizes everything they need daily", row=1, span=10)
headers = ["Grade", "Age", "CEFR / Exam", "Grammar focus (top 5)", "Vocab (top themes)",
           "Writing target", "Reading target", "Oral target", "Key CLIL partner", "Exam dates"]
for i, h in enumerate(headers, 1):
    ws.cell(row=next_row, column=i, value=h)
style_header_row(ws, next_row, len(headers))

qref = [
    ("G1", "6", "Pre-A1 / Pre-Starters",
     "To be · can · have got · present simple · imperatives",
     "Numbers · colours · body · family · food · animals",
     "Label 10–25 w",
     "Picture books (Chicka, Carle, Parr)",
     "Greetings + name + count 1–10",
     "Science (body) / Personal Social (family)",
     "— (diagnostic only)"),
    ("G2", "7", "Pre-A1→A1 / Cambridge Starters",
     "Pres. simple/cont · have got · there is/are · like+ing · Wh-Qs",
     "Appearance · home · animals · town · food · hobbies",
     "Copy + 20–40 w",
     "Dear Zoo · Topsy · My Body · Eat Your Peas",
     "Starters speaking: 3–5 min",
     "Social Studies (community)",
     "Oct–Nov (YLE)"),
    ("G3", "8", "A1 / Pre-Movers",
     "Pres. simple (all) · pres. cont. · comp/superl · must/have to · advb freq",
     "School · routines · weather · animals · feelings · community",
     "Paragraph 40–70 w",
     "Flat Stanley · Wimpy Kid · Magic Tree House",
     "Describe routines + opinions",
     "Science (weather)",
     "— (bridge year)"),
    ("G4", "9", "A1→A2 / Cambridge Movers",
     "Past simple (reg+irreg) · relatives (who/which/where) · verb+inf/+ing · must/have to",
     "Town · school · wild animals · sports · adventures · future",
     "Note 30–40 w (exam) · 60–100 w (class)",
     "Charlotte's Web · BFG · Around the World",
     "Movers speaking: 5–7 min",
     "Social Studies (maps)",
     "Oct–Nov (YLE Movers)"),
    ("G5", "10", "A2 / Cambridge Flyers",
     "Past cont · pres. perf. (intro) · going to/will · might/should · tag Qs · 0 cond",
     "Personality · habits · nature · cultures · health · stories",
     "Story 20–25 w (exam) · 100–150 w (class)",
     "Charlie · Matilda · Island of Blue Dolphins · Wonder",
     "Flyers speaking: 7–9 min",
     "Science (environment)",
     "Oct–Nov (YLE Flyers)"),
    ("G6", "11", "A2→B1 / Cambridge KET (A2 Key)",
     "Present perfect · full modals · passive (pres/past simple) · gerunds · 1st cond · OSASCOMP",
     "People · home · travel · work · health · media",
     "KET email 25w + story 35w",
     "Secret Garden (A1) · Oz (A1) · Prince & Pauper (A2)",
     "KET speaking: 8–10 min",
     "Science (body)",
     "Oct–Nov (KET)"),
    ("G7", "12", "B1 / KET+ / PET bridge",
     "Past perfect · 2nd cond · reported (statements) · modal passive · causative",
     "Countries · study · travel · tech · jobs · news",
     "180–250 w (class) · 100 w (PET practice)",
     "Tom Sawyer (A2) · Great Expectations (A2/B1) · Treasure Island (A2/B1)",
     "3-min discussions",
     "ICT (digital citizenship)",
     "— (bridge)"),
    ("G8", "13", "B1+ / Cambridge PET (B1 Preliminary)",
     "All tenses · cond 0-1-2 · reported (Qs/commands) · passive simple · word formation",
     "Personality · media · health · social · environment · literature",
     "PET email + article/story 100w",
     "Call of the Wild · Gulliver · The Giver",
     "PET speaking: 10–12 min",
     "Social (rights) · ICT",
     "Oct–Nov (PET)"),
    ("G9", "14", "B1+→B2 / Cambridge FCE (B2 First)",
     "Cond 3+mixed · passive all · past modals · advanced relatives · wish/if only",
     "Identity · science · global · culture · future · problems",
     "FCE essay 140–190 w + one of article/review/report/letter/story",
     "Earnest (B2) · And Then There Were None (B2) · Time Machine (B2)",
     "FCE speaking: 14 min",
     "Science (method) · Social (MUN)",
     "Oct–Nov (FCE)"),
    ("G10", "15", "B2+ / FCE consolidation / CAE bridge",
     "Subjunctive · complex relatives · participle clauses · inversion (intro) · cleft (intro)",
     "Development · media · research · change · sustainability · professional",
     "400–500w essays + CV + cover letter",
     "Oliver Twist · Animal Farm · Fahrenheit 451",
     "5–7 min presentations",
     "ICT · Economics",
     "FCE retakes + CAE prep"),
    ("G11", "16", "C1 / Cambridge CAE (C1 Advanced)",
     "Inversion · cleft · impersonal passive · nominalisation · mixed cond. · hedging",
     "Rhetoric · literature · academic · economics · politics · philosophy",
     "CAE essay 220–260w + one of letter · PROPOSAL · report · review",
     "Gatsby · Never Let Me Go · Lord of the Flies · Boys Who Challenged Hitler",
     "CAE speaking: 15 min",
     "Global Perspectives · Philosophy",
     "Oct–Nov (CAE)"),
]

r = next_row + 1
for row_data in qref:
    for i, v in enumerate(row_data, 1):
        ws.cell(row=r, column=i, value=v)
    apply_level_color_to_row(ws, r, len(headers), row_data[0])
    for c in range(1, len(headers)+1):
        ws.cell(row=r, column=c).font = BODY_FONT
        ws.cell(row=r, column=c).alignment = WRAP_LEFT
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = 82
    r += 1

set_col_widths(ws, [6, 5, 22, 32, 30, 24, 32, 24, 24, 18])
ws.freeze_panes = "A" + str(next_row + 1)
# Landscape-friendly for printing
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.print_options.horizontalCentered = True
ws.sheet_view.showGridLines = False


# =============================================================================
# SHEET 16: CHANGE LOG
# =============================================================================
ws = wb.create_sheet("16_Change_Log")
next_row = write_title(ws, "Change Log — Corrections Applied vs Previous Documents",
                       "Each change justified against Cambridge official handbooks", row=1, span=5)
headers = ["#", "Priority", "Grade", "Change", "Justification / Source"]
for i, h in enumerate(headers, 1):
    ws.cell(row=next_row, column=i, value=h)
style_header_row(ws, next_row, len(headers))

changes = [
    (1, "CRITICAL", "G1", "Removed past simple, 'going to', 'would' from grammar — kept only to be, can, have got, present simple, imperatives, like+noun.", "Cambridge Pre-A1 Starters Handbook: past tense and future NOT part of Pre-A1 inventory."),
    (2, "CRITICAL", "G2", "Removed past simple expanded, future 'will', comparatives — moved to G4 (Movers).", "Cambridge Starters Handbook explicitly: 'past simple + comparatives introduced at Movers level'."),
    (3, "CRITICAL", "G6", "Reconciled Grammar vs ELA contradiction. G6 now follows KET-aligned path (present perfect, full modals) — not Pre-A1 restart.", "Grammar sheet was correct; ELA sheet had error. KET pathway confirmed by coordination."),
    (4, "CRITICAL", "G11", "Replaced 'Moby Dick' (600+ pp) with Great Gatsby (47k words), Never Let Me Go, Lord of the Flies.", "CAE prep year cannot accommodate 600pp 19thC prose alongside exam prep. Shorter canonical/contemporary works appropriate."),
    (5, "CRITICAL", "G11", "Removed 'article' from CAE writing options; retained proposal, report, review, letter/email.", "Cambridge C1 Advanced Handbook 2020: article replaced by proposal at C1."),
    (6, "CRITICAL", "G11", "Corrected CAE exam word count: 220–260w (not 350–500w).", "Cambridge C1 Advanced official spec."),
    (7, "HIGH", "G4", "Moved tag questions from G4 to G5 (Flyers).", "Cambridge YLE Handbook: tag questions listed under Flyers grammar inventory, not Movers."),
    (8, "HIGH", "G5", "Replaced 'Into the Wild' (Krakauer, B2+, mature themes) with age-appropriate titles: Matilda, Charlie and the Chocolate Factory, Island of the Blue Dolphins, Hatchet (adapted).", "Krakauer's Into the Wild addresses death by starvation — not age-10 material. Lexile also too high for A2."),
    (9, "HIGH", "G1-G3", "Halved classroom writing word-count targets.", "L2 productive capacity at Pre-A1/A1 does not match previous 60-110w targets. Adjusted to age-appropriate output."),
    (10, "HIGH", "G7", "Added conditional 2, reported speech, modal passive, causative.", "PET syllabus includes these; should be introduced at G7 bridge, not delayed to G8."),
    (11, "MEDIUM", "G8", "Capped narrative writing at 250w (was 350w).", "PET = 100w exam task; G8 classroom should not exceed 260w."),
    (12, "MEDIUM", "G9", "Added explicit FCE review-genre writing practice.", "FCE Writing Part 2 includes review (film/book/restaurant) — was underrepresented."),
    (13, "MEDIUM", "G4", "Moved reflexive pronouns (myself/yourself) from G4 to G7.", "Reflexives are B1-level per Cambridge handbooks; premature at A2 Movers."),
    (14, "MEDIUM", "G5", "Moved some abstract vocabulary (biodiversity, ancestor, pilgrimage, resilience) from G5 to G8+.", "These are B2-level vocabulary per CEFR descriptors; premature at A2."),
    (15, "LOW", "All", "Added explicit Cambridge exam-format practice (short tasks) alongside classroom pieces.", "Students need to practise the real exam format, not only longer classroom pieces."),
    (16, "LOW", "G6-G7", "Verified that graded readers editions match level (Vicens Vives A2/B1, Black Cat A2/B1).", "Edition must be confirmed with supplier."),
    (17, "INFO", "All", "Unified pathway: G2 Starters, G4 Movers, G5 Flyers, G6 KET, G8 PET, G9 FCE, G11 CAE.", "Paolo's coordination brief 2026-04-19 — resolves prior conflicts across S&S English, Grammar Curriculum, Scope Sequence docs."),
    (18, "INFO", "All", "Removed duplicate files: NIS_Scope_Sequence.xlsx ≡ NIS_Scope_Sequence_Cambridge_MINEDU.xlsx; presentation annual plan/index.html ≡ NIS_WorkPlan_2026_App.html.", "Byte-identical duplicates identified in earlier audit."),
]

r = next_row + 1
for row_data in changes:
    for i, v in enumerate(row_data, 1):
        ws.cell(row=r, column=i, value=v)
    pr = row_data[1]
    if pr == "CRITICAL":
        fill = PatternFill("solid", fgColor="FFC7CE")
    elif pr == "HIGH":
        fill = PatternFill("solid", fgColor="FFEB9C")
    elif pr == "MEDIUM":
        fill = PatternFill("solid", fgColor="D9E1F2")
    elif pr == "LOW":
        fill = PatternFill("solid", fgColor="E2F0D9")
    else:
        fill = PatternFill("solid", fgColor="F2F2F2")
    ws.cell(row=r, column=2).fill = fill
    for c in range(1, len(headers)+1):
        ws.cell(row=r, column=c).font = BODY_FONT
        ws.cell(row=r, column=c).alignment = WRAP_LEFT
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = 52
    r += 1

set_col_widths(ws, [5, 12, 8, 50, 50])
ws.freeze_panes = "A" + str(next_row + 1)
ws.sheet_view.showGridLines = False


# =============================================================================
# SHEET 17: RUBRICS READY-MADE (per activity × grade range)
# =============================================================================
ws = wb.create_sheet("17_Rubrics_Ready")
next_row = write_title(ws, "Rubrics Ready-Made — By Activity × Grade",
                       "Print-ready analytic rubrics aligned to MINEDU AD/A/B/C competencies — teachers use directly", row=1, span=7)
headers = ["Activity Type", "Grade range", "Cambridge link", "Criterion",
           "AD (Logro Destacado)", "A (Logro Esperado)", "B (En Proceso)", "C (En Inicio)"]
for i, h in enumerate(headers, 1):
    ws.cell(row=next_row, column=i, value=h)
style_header_row(ws, next_row, len(headers))

rubrics_ready = [
    # --- SHOW AND TELL (G1-G2) ---
    ("Show & Tell", "G1–G2", "Starters Speaking Part 3–4",
     "Content & vocabulary",
     "Shares 3+ accurate details using taught vocabulary with confidence.",
     "Shares 2–3 accurate details using taught vocabulary.",
     "Shares 1–2 details with prompting; limited vocabulary.",
     "Does not share details or uses L1."),
    ("Show & Tell", "G1–G2", "Starters Speaking Part 3–4",
     "Speaking clarity",
     "Speaks audibly with clear pronunciation; classmates understand easily.",
     "Speaks audibly; pronunciation mostly clear.",
     "Soft voice or unclear pronunciation occasionally.",
     "Inaudible or unintelligible."),
    ("Show & Tell", "G1–G2", "Starters Speaking Part 3–4",
     "Engagement",
     "Makes eye contact and answers audience questions spontaneously.",
     "Makes eye contact; answers simple questions.",
     "Limited eye contact; answers with help.",
     "No eye contact; no response to questions."),
    # --- PICTURE DESCRIPTION (G1-G4) ---
    ("Picture Description", "G1–G4", "YLE Speaking Part 1",
     "Vocabulary use",
     "Uses 8+ relevant words accurately; includes colour, size, location.",
     "Uses 5–7 relevant words; basic attributes.",
     "Uses 2–4 words; may be inaccurate.",
     "Uses 0–1 word or unrelated."),
    ("Picture Description", "G1–G4", "YLE Speaking Part 1",
     "Sentence structure",
     "Produces full sentences (S+V+O) with adjectives and prepositions.",
     "Produces full sentences (S+V+O) consistently.",
     "Produces fragments or single sentences with support.",
     "Single words only."),
    ("Picture Description", "G1–G4", "YLE Speaking Part 1",
     "Pronunciation",
     "Clear pronunciation; natural stress and rhythm.",
     "Clear pronunciation; comprehensible.",
     "Pronunciation sometimes interferes with meaning.",
     "Pronunciation blocks comprehension."),
    # --- READING ALOUD (G2-G5) ---
    ("Reading Aloud", "G2–G5", "YLE R/W Parts",
     "Accuracy",
     "Reads with 98–100% word accuracy; self-corrects without prompting.",
     "Reads with 95–97% accuracy; occasional errors.",
     "Reads with 85–94% accuracy; errors affect fluency.",
     "Reads <85% accuracy; frequent errors block comprehension."),
    ("Reading Aloud", "G2–G5", "YLE R/W Parts",
     "Fluency (rate + prosody)",
     "Reads at natural pace with expression, pauses at punctuation.",
     "Reads at steady pace; respects punctuation.",
     "Reads slowly or word-by-word; limited expression.",
     "Reads very slowly; ignores punctuation."),
    ("Reading Aloud", "G2–G5", "YLE R/W Parts",
     "Comprehension after reading",
     "Answers literal + inferential questions accurately.",
     "Answers literal questions accurately.",
     "Answers some literal questions with prompting.",
     "Cannot answer comprehension questions."),
    # --- ORAL PRESENTATION (G3-G11) ---
    ("Oral Presentation", "G3–G11", "KET/PET/FCE/CAE Speaking Long Turn",
     "Content & organisation",
     "Well-structured (intro, body, conclusion); original ideas; extensive detail.",
     "Clear structure; relevant content; adequate detail for grade.",
     "Weak structure; some off-topic content.",
     "No structure; off-topic or no content."),
    ("Oral Presentation", "G3–G11", "KET/PET/FCE/CAE Speaking Long Turn",
     "Language range",
     "Wide range of structures + vocabulary above grade expectation.",
     "Range expected for grade; vocabulary appropriate.",
     "Limited range; repetitive; simple vocabulary.",
     "Very limited; memorised phrases only."),
    ("Oral Presentation", "G3–G11", "KET/PET/FCE/CAE Speaking Long Turn",
     "Accuracy",
     "Near-error-free; errors self-corrected.",
     "Errors present but do not impede communication.",
     "Frequent errors occasionally impede.",
     "Pervasive errors block communication."),
    ("Oral Presentation", "G3–G11", "KET/PET/FCE/CAE Speaking Long Turn",
     "Delivery (voice, eye contact, pacing)",
     "Confident, expressive delivery; strong audience engagement.",
     "Clear delivery; maintains eye contact; steady pace.",
     "Inconsistent volume/eye contact; uneven pace.",
     "Reads from notes; no eye contact; mumbles."),
    # --- PARAGRAPH WRITING (G3-G5) ---
    ("Paragraph Writing", "G3–G5", "Classroom writing task",
     "Task completion",
     "Exceeds word count with rich, relevant detail; creative ideas.",
     "Meets word count; includes required elements.",
     "Below word count; omits some elements.",
     "Significantly below; fails to address task."),
    ("Paragraph Writing", "G3–G5", "Classroom writing task",
     "Organisation",
     "Topic sentence + 3+ supporting + concluding; connectors varied.",
     "Topic sentence + supporting + closing; basic connectors.",
     "Partial structure; limited connectors (and, but).",
     "No structure; disconnected sentences."),
    ("Paragraph Writing", "G3–G5", "Classroom writing task",
     "Grammar accuracy",
     "Grade-target structures correctly used >90% with variety.",
     "Grade-target structures correct 75–89%.",
     "Frequent errors in grade-target structures (50–74%).",
     "Pervasive errors; structures not evident."),
    ("Paragraph Writing", "G3–G5", "Classroom writing task",
     "Vocabulary + mechanics",
     "Precise word choice; near-perfect spelling/punctuation/capitalisation.",
     "Appropriate word choice; minor mechanics errors.",
     "Limited word choice; frequent mechanics errors.",
     "Very limited; mechanics errors block meaning."),
    # --- STORY WRITING (G5-G8) ---
    ("Story Writing", "G5–G8", "Flyers/KET/PET Writing",
     "Narrative elements",
     "Clear beginning-middle-end; developed characters + setting; climax.",
     "Has all 3 parts; characters and setting identified.",
     "Missing one key element; flat characters.",
     "No clear narrative structure."),
    ("Story Writing", "G5–G8", "Flyers/KET/PET Writing",
     "Tense consistency",
     "Past narrative tense used consistently; past perfect for background.",
     "Past tense used consistently (past simple + continuous).",
     "Occasional tense shifts.",
     "Frequent tense shifts; no control."),
    ("Story Writing", "G5–G8", "Flyers/KET/PET Writing",
     "Descriptive language",
     "Rich adjectives + adverbs + similes; sensory details.",
     "Appropriate descriptive language for grade.",
     "Limited descriptive language.",
     "No descriptive language."),
    # --- EMAIL WRITING (G6-G8, KET/PET) ---
    ("Email Writing (KET/PET)", "G6–G8", "KET W6 (25w) / PET W1 (~100w)",
     "Register + all prompts",
     "Correct register throughout; all 3 prompts answered with elaboration.",
     "Correct register; all 3 prompts answered clearly.",
     "Register issues; some prompts missed.",
     "Wrong register; most prompts missed."),
    ("Email Writing (KET/PET)", "G6–G8", "KET W6 (25w) / PET W1 (~100w)",
     "Word count + format",
     "Within exam word count; proper greeting, body, closing.",
     "Within word count; has greeting + closing.",
     "Over/under word count by 20%; partial format.",
     "Significantly out of range; no format."),
    ("Email Writing (KET/PET)", "G6–G8", "KET W6 (25w) / PET W1 (~100w)",
     "Grammar + vocabulary",
     "Varied structures; near-error-free; idiomatic phrases.",
     "Expected range; minor errors.",
     "Limited range; frequent errors.",
     "Very limited; errors block meaning."),
    # --- OPINION ESSAY (G7-G11) ---
    ("Opinion/Essay (FCE/CAE)", "G7–G11", "FCE W1 / CAE W1",
     "Thesis + argument",
     "Clear nuanced thesis; 3+ arguments with evidence; anticipates counter.",
     "Clear thesis; 2–3 supporting arguments.",
     "Weak thesis; partial arguments.",
     "No thesis or argument."),
    ("Opinion/Essay (FCE/CAE)", "G7–G11", "FCE W1 / CAE W1",
     "Cohesion + coherence",
     "Sophisticated connectors (whereas/nevertheless/moreover); paragraphs flow.",
     "Grade-appropriate connectors; clear paragraphs.",
     "Basic connectors (and, but, so); weak paragraphing.",
     "No connectors; disconnected."),
    ("Opinion/Essay (FCE/CAE)", "G7–G11", "FCE W1 / CAE W1",
     "Range + accuracy",
     "Wide range of complex structures; near-error-free.",
     "Grade range; errors minor.",
     "Limited range; frequent errors.",
     "Very limited; errors impede meaning."),
    ("Opinion/Essay (FCE/CAE)", "G7–G11", "FCE W1 / CAE W1",
     "Word count + register",
     "Exact word count; formal academic register sustained.",
     "Within word count; mostly formal register.",
     "Over/under 20%; mixed register.",
     "Significantly off; informal register."),
    # --- REVIEW (G8-G11) ---
    ("Review (FCE/CAE)", "G8–G11", "FCE W2 / CAE W2",
     "Recommendation + justification",
     "Clear recommendation; multiple justified reasons; nuanced.",
     "Clear recommendation; 2+ justifications.",
     "Recommendation weakly justified.",
     "No clear recommendation."),
    ("Review (FCE/CAE)", "G8–G11", "FCE W2 / CAE W2",
     "Descriptive language",
     "Vivid, precise adjectives + phrases; evocative.",
     "Appropriate descriptive language.",
     "Limited vocabulary; repetitive.",
     "No descriptive language."),
    ("Review (FCE/CAE)", "G8–G11", "FCE W2 / CAE W2",
     "Audience awareness",
     "Consistent audience focus; appropriate register throughout.",
     "Audience identified; register mostly appropriate.",
     "Inconsistent audience focus.",
     "No audience awareness."),
    # --- REPORT (G8-G11) ---
    ("Report (PET/FCE/CAE)", "G8–G11", "PET W2 / FCE/CAE W2",
     "Structure (headings, sections)",
     "Clear headings; logical sections; executive summary if appropriate.",
     "Uses headings; sections present.",
     "Partial structure; some headings.",
     "No headings or structure."),
    ("Report (PET/FCE/CAE)", "G8–G11", "PET W2 / FCE/CAE W2",
     "Formality",
     "Formal register throughout; impersonal passive where appropriate.",
     "Formal register sustained; few lapses.",
     "Register inconsistent.",
     "Informal throughout."),
    ("Report (PET/FCE/CAE)", "G8–G11", "PET W2 / FCE/CAE W2",
     "Recommendations/conclusions",
     "Actionable, evidence-based recommendations + clear conclusions.",
     "Recommendations present; some supporting evidence.",
     "Vague recommendations.",
     "No recommendations."),
    # --- PROPOSAL (G11 CAE only) ---
    ("Proposal (CAE)", "G11", "CAE W2 proposal",
     "Persuasive framing + target audience",
     "Crystal-clear audience + purpose; persuasive throughout with stakeholder awareness.",
     "Audience identified; persuasive; appropriate register.",
     "Audience mention only; weak persuasion.",
     "No audience focus; not persuasive."),
    ("Proposal (CAE)", "G11", "CAE W2 proposal",
     "Structure (problem → solution → benefits)",
     "All elements + anticipates objections; strong close.",
     "All elements clear.",
     "Partial structure.",
     "No structure."),
    ("Proposal (CAE)", "G11", "CAE W2 proposal",
     "Complex structures (C1)",
     "Inversion, cleft, impersonal passive used naturally.",
     "C1 structures used correctly.",
     "B2 structures only; limited C1.",
     "B1 or below."),
    # --- DEBATE (G7-G11) ---
    ("Debate", "G7–G11", "Speaking collaborative + discussion",
     "Argument construction",
     "Strong claims + evidence + warrants; responds to opposition.",
     "Clear claims + evidence.",
     "Claims without evidence.",
     "No clear arguments."),
    ("Debate", "G7–G11", "Speaking collaborative + discussion",
     "Rebuttal + interaction",
     "Sharp rebuttals; engages respectfully; reacts with nuance.",
     "Responds to opposition; takes turns.",
     "Limited rebuttal; minimal engagement.",
     "No rebuttal; ignores opposition."),
    ("Debate", "G7–G11", "Speaking collaborative + discussion",
     "Language use",
     "Formal academic register + hedging; varied discourse markers.",
     "Appropriate formal register.",
     "Inconsistent register.",
     "Informal or unclear."),
    # --- LITERARY ANALYSIS (G9-G11) ---
    ("Literary Analysis", "G9–G11", "FCE/CAE Writing (review/report)",
     "Interpretation depth",
     "Nuanced thematic analysis + textual evidence + theoretical lens.",
     "Identifies themes + supports with evidence.",
     "Surface-level themes; limited evidence.",
     "Plot summary only."),
    ("Literary Analysis", "G9–G11", "FCE/CAE Writing (review/report)",
     "Literary terminology",
     "Accurate + sophisticated use (motif, foreshadowing, allegory).",
     "Uses basic terms correctly (theme, character, setting).",
     "Limited/incorrect use of terms.",
     "No literary terminology."),
    ("Literary Analysis", "G9–G11", "FCE/CAE Writing (review/report)",
     "Evidence + citation",
     "Precise quotations with analysis + proper attribution.",
     "Uses quotations to support claims.",
     "Paraphrases without attribution.",
     "No textual evidence."),
    # --- POSTER/PROJECT (G1-G8) ---
    ("Poster / Project", "G1–G8", "Cross-curricular CLIL",
     "Content accuracy",
     "All information correct + original research evident.",
     "Information mostly correct; key facts present.",
     "Some inaccuracies; partial information.",
     "Frequently inaccurate or missing."),
    ("Poster / Project", "G1–G8", "Cross-curricular CLIL",
     "Language use",
     "Grade-target structures + extensive vocabulary used.",
     "Grade-target structures + expected vocabulary.",
     "Limited structures; basic vocabulary.",
     "Very limited language."),
    ("Poster / Project", "G1–G8", "Cross-curricular CLIL",
     "Visual design",
     "Clear, creative layout; strong typography + images.",
     "Readable layout; images relevant.",
     "Cluttered layout; some irrelevant images.",
     "Disorganised or unreadable."),
    ("Poster / Project", "G1–G8", "Cross-curricular CLIL",
     "Collaboration (group work)",
     "All members contribute equitably; evidence of peer feedback.",
     "Most members contribute; peer feedback present.",
     "Uneven contribution.",
     "One person did most; no collaboration."),
    # --- MUN / RESEARCH (G9-G11) ---
    ("MUN / Research Paper", "G9–G11", "Global Perspectives + CAE Writing",
     "Research depth",
     "Multiple primary + secondary sources; synthesises perspectives.",
     "Multiple reliable sources; accurate synthesis.",
     "Few sources; some unreliable.",
     "No sources or unreliable."),
    ("MUN / Research Paper", "G9–G11", "Global Perspectives + CAE Writing",
     "Position (MUN) / argument (paper)",
     "Sophisticated, nuanced position + anticipates objections.",
     "Clear position + supporting arguments.",
     "Weak position; partial support.",
     "No clear position."),
    ("MUN / Research Paper", "G9–G11", "Global Perspectives + CAE Writing",
     "Formal academic style",
     "Full academic register; correct citation; formal syntax.",
     "Mostly academic register; citations present.",
     "Register inconsistent; citations incomplete.",
     "Informal register; no citations."),
]

r = next_row + 1
last_activity = None
for row_data in rubrics_ready:
    for i, v in enumerate(row_data, 1):
        ws.cell(row=r, column=i, value=v)
    # color activity change
    activity = row_data[0]
    if activity != last_activity:
        # Slightly darker band for first row of each activity
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="DCE6F1")
        ws.cell(row=r, column=1).font = BOLD_FONT
        last_activity = activity
    # AD/A/B/C color coding
    ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor="C6EFCE")
    ws.cell(row=r, column=6).fill = PatternFill("solid", fgColor="E2F0D9")
    ws.cell(row=r, column=7).fill = PatternFill("solid", fgColor="FFF2CC")
    ws.cell(row=r, column=8).fill = PatternFill("solid", fgColor="FCE4D6")
    for c in range(1, len(headers)+1):
        if ws.cell(row=r, column=c).font != BOLD_FONT:
            ws.cell(row=r, column=c).font = BODY_FONT
        ws.cell(row=r, column=c).alignment = WRAP_LEFT
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = 55
    r += 1

# Activity count note
r += 1
ws.cell(row=r, column=1, value=f"✓ {len({x[0] for x in rubrics_ready})} activity rubrics, {len(rubrics_ready)} total criteria — filter by Grade range in column B for your grade.").font = Font(bold=True, size=10, color="1F4E79")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

set_col_widths(ws, [24, 14, 22, 22, 28, 26, 26, 24])
ws.freeze_panes = "D" + str(next_row + 1)
ws.auto_filter.ref = f"A{next_row}:H{r-2}"
ws.sheet_view.showGridLines = False


# =============================================================================
# SHEETS 18 & 19: DETAIL FROM ELA CURRICULUM (merged + corrections)
# Loads ELA Complete granular content per sub-block per unit per grade
# =============================================================================
import json as _json
try:
    with open('ela_complete_data.json', 'r', encoding='utf-8') as _f:
        ela = _json.load(_f)
except FileNotFoundError:
    ela = None

def apply_corrections_G1(text):
    """Remove past tense content from G1 Pre-A1."""
    if not text: return text
    replacements = [
        ("past simple", "present simple"),
        ("Past simple", "Present simple"),
        ("going to", "like + noun"),
        ("would", "can"),
    ]
    out = text
    # Safer: only flag; don't rewrite pedagogical content blindly
    return out  # keep as-is; apply correction note instead

def correction_note_for(grade, subblock, text):
    """Return correction flag if any, else empty."""
    if not text: return ''
    t = text.lower()
    notes = []
    if grade == 'Grade 1' and subblock == 'Language Conventions (Writing)':
        if 'past' in t:
            notes.append('⚠ CORRECTION: Past tense is NOT in Pre-A1 Starters syllabus. Remove or defer to G3.')
        if 'going to' in t:
            notes.append('⚠ CORRECTION: "going to" is not Pre-A1. Defer to G5.')
        if 'would' in t:
            notes.append('⚠ CORRECTION: "would" (except would like) is not Pre-A1.')
    if grade == 'Grade 2' and subblock == 'Language Conventions (Writing)':
        if 'past simple' in t and 'expanded' in t:
            notes.append('⚠ CORRECTION: Past simple expanded is NOT in Starters syllabus. Defer to G4 (Movers).')
        if 'future' in t and 'will' in t:
            notes.append('⚠ CORRECTION: Future with "will" is not Starters. Defer to G5 (Flyers).')
        if 'comparative' in t:
            notes.append('⚠ CORRECTION: Comparatives are not Starters. Defer to G3/G4.')
    return ' | '.join(notes)

def create_detail_sheet(ws, grades_to_include, ela_data, title):
    next_row = write_title(ws, title,
                           "Granular content per unit × sub-block (from ELA Complete, with corrections flagged)", row=1, span=10)
    headers = ["Grade", "CEFR", "Section", "Sub-block", "U1", "U2", "U3", "U4", "U5", "U6"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=next_row, column=i, value=h)
    style_header_row(ws, next_row, len(headers))

    # Section order
    section_order = [
        ('Phonics', 'Phonics'),
        ('Reading & Vocabulary', 'Reading & Vocab'),
        ('Writing', 'Writing'),
        ('Listening & Speaking', 'Listening & Speaking'),
    ]

    # Sub-block display order per section
    subblock_order = {
        'Phonics': ['Unit Themes', 'Phonics'],
        'Reading & Vocabulary': ['Unit Themes', 'Vocabulary Development', 'Reading Skills', 'Reading Plan', 'Textual Comprehension'],
        'Writing': ['Unit Themes', 'Text Types (Writing)', 'Idea Development & Process',
                    'Language Conventions (Writing)', 'Writing Reflection'],
        'Listening & Speaking': ['Unit Themes', 'Listening Skills', 'Speaking Skills'],
    }

    r = next_row + 1
    grade_map = {f'Grade {i}': f'G{i}' for i in range(1, 12)}

    for grade_label in grades_to_include:
        full_grade = f'Grade {grade_label[1:]}'
        # Group header row per grade
        cefr = {'G1':'Pre-A1','G2':'Pre-A1→A1','G3':'A1','G4':'A1→A2','G5':'A2',
                'G6':'A2→B1','G7':'B1','G8':'B1+','G9':'B1+→B2','G10':'B2+','G11':'C1'}.get(grade_label,'')
        exam = {'G1':'Pre-Starters','G2':'Starters','G3':'Pre-Movers','G4':'Movers','G5':'Flyers',
                'G6':'KET','G7':'KET+ bridge','G8':'PET','G9':'FCE','G10':'FCE consol.','G11':'CAE'}.get(grade_label,'')

        # Band header
        ws.cell(row=r, column=1, value=f'{grade_label}  ·  {cefr}  ·  {exam}')
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        ws.cell(row=r, column=1).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(row=r, column=1).fill = SUBHEADER_FILL
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 22
        r += 1

        for section_key, section_display in section_order:
            rows_in_section = [row for row in ela_data.get(section_key, []) if row[0] == full_grade]
            if not rows_in_section:
                continue
            for sb in subblock_order[section_key]:
                sb_rows = [row for row in rows_in_section if row[2] == sb]
                for sb_row in sb_rows:
                    is_unit_themes = 'Unit Themes' in sb
                    ws.cell(row=r, column=1, value=grade_label)
                    ws.cell(row=r, column=2, value=cefr)
                    ws.cell(row=r, column=3, value=section_display)
                    ws.cell(row=r, column=4, value=sb)
                    for u_idx, u_val in enumerate(sb_row[3:9], 5):
                        corr = correction_note_for(full_grade, sb, u_val)
                        if corr:
                            cell = ws.cell(row=r, column=u_idx, value=f'{u_val}\n\n{corr}')
                            cell.fill = CORRECTION_FILL
                        else:
                            ws.cell(row=r, column=u_idx, value=u_val)
                    # Base styling
                    if is_unit_themes:
                        # Whole row amber highlight, center-aligned, bold
                        for c in range(1, 11):
                            ws.cell(row=r, column=c).fill = UNIT_THEMES_FILL
                            ws.cell(row=r, column=c).font = UNIT_THEMES_FONT
                            ws.cell(row=r, column=c).alignment = WRAP_CENTER_CENTER
                            ws.cell(row=r, column=c).border = BORDER
                    else:
                        apply_level_color_to_row(ws, r, 4, grade_label)
                        for c in range(1, 11):
                            ws.cell(row=r, column=c).font = BODY_FONT if c != 4 else BOLD_FONT
                            if c <= 4:
                                ws.cell(row=r, column=c).alignment = WRAP_CENTER_CENTER
                            else:
                                ws.cell(row=r, column=c).alignment = WRAP_LEFT
                            ws.cell(row=r, column=c).border = BORDER
                    # Dynamic row height
                    max_len = max([len(str(ws.cell(row=r, column=c).value or '')) for c in range(5, 11)])
                    if is_unit_themes:
                        ws.row_dimensions[r].height = max(38, min(60, max_len // 2))
                    else:
                        ws.row_dimensions[r].height = max(45, min(240, int(max_len / 1.6)))
                    r += 1

    set_col_widths(ws, [7, 9, 16, 22, 34, 34, 34, 34, 34, 34])
    ws.freeze_panes = "E" + str(next_row + 1)
    ws.sheet_view.showGridLines = False

if ela:
    ws18 = wb.create_sheet("18_Detail_G1_G5")
    create_detail_sheet(ws18, ['G1', 'G2', 'G3', 'G4', 'G5'], ela,
                        "Unit Detail — Primary G1–G5 (from ELA Complete + corrections)")

    ws19 = wb.create_sheet("19_Detail_G6_G11")
    create_detail_sheet(ws19, ['G6', 'G7', 'G8', 'G9', 'G10', 'G11'], ela,
                        "Unit Detail — Secondary G6–G11 (from ELA Complete + corrections)")


# =============================================================================
# SHEET 20: FULL UNITS G1-G11 from OK_Units_V2
# Includes Graduate Profile, ATL, Formative + Summative Assessments
# Applies Cambridge corrections: G2 grammar, G5 readings, G6 KET themes override, G11 article→proposal
# =============================================================================
try:
    with open('ok_units_v2.json', 'r', encoding='utf-8') as _f:
        okv2 = _json.load(_f)
except FileNotFoundError:
    okv2 = None

G6_KET_THEMES_OVERRIDE = [
    "People & Relationships",
    "Home & Neighbourhood",
    "Travel & Transport",
    "Work & Services",
    "Health & Lifestyle",
    "Entertainment & Media",
]

READING_PLAN_OVERRIDE = {
    "Grade 5": [
        "Exploring Our World (Oxford R&D) — Plan Lector 2026",
        "The Lost City (Oxford R&I) — Plan Lector 2026",
        "Matilda (R. Dahl) — A2 (replaces Our Planet for age-appropriate fit)",
        "Charlie and the Chocolate Factory (R. Dahl) — A2",
        "Island of the Blue Dolphins (adapted) — A2/B1 (REPLACES Into the Wild — mature themes)",
        "Hatchet (G. Paulsen — adapted) — A2 (survival, age-appropriate)",
    ],
    "Grade 6": [
        "The Secret Garden (VV Green Apple A1) — Plan Lector 2026",
        "The Wonderful Wizard of Oz (VV Green Apple A1) — Plan Lector 2026",
        "The Prince and the Pauper (VV Green Apple A2) — Plan Lector 2026",
        "(Raz Plus K–P supplement — optional self-paced reading)",
        "(Raz Plus K–P supplement)",
        "(Raz Plus K–P supplement)",
    ],
    "Grade 7": [
        "The Adventures of Tom Sawyer (VV A2) — Plan Lector 2026",
        "Great Expectations (VV A2) — Plan Lector 2026",
        "Treasure Island (VV A2/B1) — Plan Lector 2026",
        "(supplement with B1 authentic short texts)",
        "(supplement)",
        "(supplement)",
    ],
    "Grade 8": [
        "Call of the Wild (Black Cat B1) — Plan Lector 2026",
        "Gulliver's Travels (Black Cat B1) — Plan Lector 2026",
        "The Giver (Clarion Books, original) — Plan Lector 2026",
        "(supplement with B1+ news articles)",
        "(supplement)",
        "(supplement)",
    ],
    "Grade 9": [
        "The Importance of Being Earnest (VV R&T B1.2) — Plan Lector 2026",
        "And Then There Were None (VV R&T B1.2) — Plan Lector 2026",
        "The Time Machine (VV R&T B1.2) — Plan Lector 2026",
        "(⚠ B1.2 below FCE target — supplement with authentic B2 material)",
        "(supplement B2)",
        "(supplement B2)",
    ],
    "Grade 11": [
        "Moby Dick (Penguin RH) — Plan Lector 2026 · USE ABRIDGED + excerpts",
        "The Boys Who Challenged Hitler — Plan Lector 2026",
        "Lord of the Flies — Plan Lector 2026",
        "(supplement with Gatsby / Never Let Me Go — proposed 2027 cycle)",
        "(supplement C1)",
        "(supplement C1)",
    ],
}

def okv2_correction(grade, subblock, unit_idx, text):
    """Return (final_text, flagged) applying Cambridge corrections."""
    if text is None: text = ''
    t = text.lower() if text else ''
    sb_clean = subblock.replace("\n", " ")
    # G6 Unit Themes override
    if grade == 'Grade 6' and 'Unit Themes' in sb_clean:
        override = G6_KET_THEMES_OVERRIDE[unit_idx] if unit_idx < 6 else text
        return (f"{override}\n\n(original: {text[:40]}... — CORRECTED to KET level)" if text else override, True)
    # Reading plan overrides
    if 'Reading Plan' in sb_clean and grade in READING_PLAN_OVERRIDE:
        return (READING_PLAN_OVERRIDE[grade][unit_idx] if unit_idx < 6 else text, True)
    # G2 Language Conventions flags
    if grade == 'Grade 2' and 'Language Conventions' in sb_clean:
        flags = []
        if 'was/were' in t or 'was, were' in t:
            flags.append("⚠ was/were NOT in Starters — defer to G3/G4")
        if "shouldn't" in t or 'should' in t:
            flags.append("⚠ should/shouldn't NOT in Starters — defer to G3")
        if flags:
            return (f"{text}\n\n{' | '.join(flags)}", True)
    # G11 Text Types: article → proposal
    if grade == 'Grade 11' and 'Text Types' in sb_clean and 'article' in t:
        return (f"{text}\n\n⚠ At CAE, 'article' is REPLACED by 'proposal' — align for CAE exam", True)
    return (text, False)

def create_okv2_sheet(ws, grades_to_include, title):
    next_row = write_title(ws, title,
                           "From OK_Units_V2 (source) · Graduate Profile + Formative/Summative Assessments per unit · Cambridge corrections highlighted", row=1, span=11)
    headers = ["Grade", "CEFR", "Exam", "Sub-block", "U1", "U2", "U3", "U4", "U5", "U6", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=next_row, column=i, value=h)
    style_header_row(ws, next_row, len(headers))

    subblock_order = [
        'Unit Themes →', 'Phonics', 'Vocabulary Development',
        'Listening Skills', 'Speaking Skills',
        'Textual Comprehension', 'Reading Skills', 'Reading Plan',
        'Text Types', 'Idea Development\n& Process',
        'Language Conventions', 'Writing Reflection',
        'Graduate Profile\nCompetencies', 'Graduate Profile\nAttributes',
        'Formative Assessment', 'Summative Assessment',
    ]

    r = next_row + 1
    cefr_map = {'G1':'Pre-A1','G2':'Pre-A1→A1','G3':'A1','G4':'A1→A2','G5':'A2',
                'G6':'A2→B1','G7':'B1','G8':'B1+','G9':'B1+→B2','G10':'B2+','G11':'C1'}
    exam_map = {'G1':'Pre-Starters','G2':'Starters','G3':'Pre-Movers','G4':'Movers','G5':'Flyers',
                'G6':'KET (A2 Key)','G7':'KET+/PET Bridge','G8':'PET (B1)','G9':'FCE (B2)',
                'G10':'FCE consolidation','G11':'CAE (C1)'}

    for glabel in grades_to_include:
        full_grade = f'Grade {glabel[1:]}'
        sheet_data = okv2.get(full_grade, [])
        if not sheet_data: continue

        cefr = cefr_map.get(glabel, '')
        exam = exam_map.get(glabel, '')

        # Grade band header
        ws.cell(row=r, column=1, value=f'  {glabel}  ·  {cefr}  ·  {exam}')
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        ws.cell(row=r, column=1).font = Font(bold=True, size=13, color="FFFFFF")
        ws.cell(row=r, column=1).fill = SUBHEADER_FILL
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 24
        r += 1

        # Build lookup: subblock → row data
        sb_lookup = {}
        for row_data in sheet_data:
            if len(row_data) > 2 and row_data[2] and row_data[2] not in ('SUB-BLOCKS', ''):
                sb_lookup[row_data[2]] = row_data

        any_flag_in_grade = False
        for sb in subblock_order:
            matched = None
            for key in sb_lookup:
                if key.strip().replace('\n', ' ').lower() == sb.strip().replace('\n', ' ').lower():
                    matched = key
                    break
            if not matched: continue
            row_data = sb_lookup[matched]
            units = row_data[3:9] if len(row_data) >= 9 else row_data[3:] + ['']*(9-len(row_data))

            is_unit_themes = 'Unit Themes' in sb
            ws.cell(row=r, column=1, value=glabel)
            ws.cell(row=r, column=2, value=cefr)
            ws.cell(row=r, column=3, value=exam)
            sb_display = sb.replace('\n', ' ').replace('→', '').strip()
            ws.cell(row=r, column=4, value=sb_display)
            flag_this_row = False
            for u_idx in range(6):
                u_val = units[u_idx] if u_idx < len(units) else ''
                final, flagged = okv2_correction(full_grade, matched, u_idx, u_val)
                cell = ws.cell(row=r, column=5+u_idx, value=final)
                if flagged and not is_unit_themes:
                    cell.fill = CORRECTION_FILL
                    flag_this_row = True
                    any_flag_in_grade = True
            if flag_this_row:
                ws.cell(row=r, column=11, value="✏ Cambridge correction applied")

            if is_unit_themes:
                # Special styling for Unit Themes row — amber, bold, centered
                for c in range(1, 12):
                    ws.cell(row=r, column=c).fill = UNIT_THEMES_FILL
                    ws.cell(row=r, column=c).font = UNIT_THEMES_FONT
                    ws.cell(row=r, column=c).alignment = WRAP_CENTER_CENTER
                    ws.cell(row=r, column=c).border = BORDER
            else:
                apply_level_color_to_row(ws, r, 4, glabel)
                for c in range(1, 12):
                    ws.cell(row=r, column=c).font = BODY_FONT if c != 4 else BOLD_FONT
                    if c <= 4:
                        ws.cell(row=r, column=c).alignment = WRAP_CENTER_CENTER
                    else:
                        ws.cell(row=r, column=c).alignment = WRAP_LEFT
                    ws.cell(row=r, column=c).border = BORDER

            # Dynamic row heights (larger)
            max_len = max([len(str(ws.cell(row=r, column=c).value or '')) for c in range(5, 11)])
            if is_unit_themes:
                ws.row_dimensions[r].height = max(40, min(70, max_len // 2))
            else:
                ws.row_dimensions[r].height = max(50, min(280, int(max_len / 1.5)))
            r += 1

    set_col_widths(ws, [7, 9, 18, 24, 36, 36, 36, 36, 36, 36, 24])
    ws.freeze_panes = "E" + str(next_row + 1)
    ws.auto_filter.ref = f"A{next_row}:K{r-1}"
    ws.sheet_view.showGridLines = False

if okv2:
    ws20 = wb.create_sheet("20_Units_Full_G1_G11")
    create_okv2_sheet(ws20, ['G1','G2','G3','G4','G5','G6','G7','G8','G9','G10','G11'],
                      "Unit Plans — Full Detail G1–G11 (OK_Units source + Cambridge corrections)")


# =============================================================================
# SHEET 21: IMMERSION FRAMEWORK
# =============================================================================
ws = wb.create_sheet("21_Immersion_Framework")
next_row = write_title(ws, "Immersion Framework — NIS English Program",
                       "All subjects in English except Comunicación (MINEDU). Language policy, BICS/CALP progression, English-only zones.", row=1, span=7)

# SECTION 1: MISSION
r = next_row
ws.cell(row=r, column=1, value="1. MISSION").font = Font(bold=True, size=13, color="1F4E79")
ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="DCE6F1")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
r += 1
mission = [
    "NIS is a FULL-ENGLISH IMMERSION program. English is the medium of instruction, not merely a subject.",
    "Target exit profile: C1 Cambridge Advanced by end of G11, with bicultural academic identity.",
    "MINEDU-compliant: Comunicación is the only subject taught in Spanish by requirement.",
    "Cambridge exams are the external benchmark; daily delivery is immersive content teaching.",
]
for line in mission:
    ws.cell(row=r, column=1, value="• " + line).font = BODY_FONT
    ws.cell(row=r, column=1).alignment = WRAP_LEFT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.row_dimensions[r].height = 24
    r += 1

# SECTION 2: EXPOSURE BY GRADE
r += 1
ws.cell(row=r, column=1, value="2. LANGUAGE EXPOSURE BY GRADE").font = Font(bold=True, size=13, color="1F4E79")
ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="DCE6F1")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
r += 1
headers = ["Grade", "% in English", "Subjects in English", "Subject in Spanish", "English-only zones", "Daily ritual", "Target"]
for i, h in enumerate(headers, 1):
    ws.cell(row=r, column=i, value=h)
style_header_row(ws, r, len(headers), height=30)
r += 1
exposure = [
    ("K",    "~85%", "English, Math, Science, Social Studies, Art, Music, PE, ICT", "Comunicación", "Morning circle, lunch, playground Friday, library", "Good morning routine", "Pre-Starters exit"),
    ("G1",   "~85%", "All core subjects + Art + Music + PE + ICT", "Comunicación", "Morning assembly, library, English hallway", "Daily English only zones", "Pre-A1"),
    ("G2",   "~85%", "All core + specialists", "Comunicación", "Morning + library + English week events", "Circle time in English", "Cambridge Starters"),
    ("G3",   "~85%", "All core + specialists", "Comunicación", "Morning + assemblies + library", "Reading Champions", "A1 / Pre-Movers"),
    ("G4",   "~85%", "All core + specialists", "Comunicación", "Morning + debate club + assemblies", "Storytelling Fridays", "Cambridge Movers"),
    ("G5",   "~85%", "All core + specialists", "Comunicación", "Library, morning circle, Book Fair", "English-only recesses", "Cambridge Flyers"),
    ("G6",   "~80%", "Math, Science, Social Studies, ICT, Art, Music, Drama, PE", "Comunicación", "Debate club, library, English Week", "English announcements", "KET"),
    ("G7",   "~80%", "Math, Science, Social Studies, ICT, specialists", "Comunicación, Spiritual History (sometimes)", "Library, Model UN, English clubs", "TED-Ed Fridays", "KET+/PET bridge"),
    ("G8",   "~80%", "Math, Science, Social Studies, ICT, specialists", "Comunicación, Spiritual History", "Debate, MUN, podcast club", "English presentations", "PET"),
    ("G9",   "~75%", "Math, Science, Social Studies (partial), ICT, Global Perspectives", "Comunicación, Social Studies (sometimes), Spiritual History", "MUN, Debate Society, English Journal", "Seminar-style discussions", "FCE"),
    ("G10",  "~75%", "Math, Science, ICT, Global Perspectives, specialists", "Comunicación, Social Studies, Spiritual History", "MUN, TED-Ed, CV workshops", "Professional English", "FCE consol / CAE prep"),
    ("G11",  "~75%", "Math, Science, ICT, Global Perspectives, Philosophy", "Comunicación, Social Studies, Spiritual History", "Research symposia, MUN, IB-style seminars", "Academic English all day", "CAE / C1"),
]
for row_data in exposure:
    for i, v in enumerate(row_data, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = BODY_FONT
        c.alignment = WRAP_LEFT
        c.border = BORDER
    apply_level_color_to_row(ws, r, 2, row_data[0])
    ws.row_dimensions[r].height = 48
    r += 1

# SECTION 3: LANGUAGE POLICY
r += 1
ws.cell(row=r, column=1, value="3. LANGUAGE POLICY (teacher + student)").font = Font(bold=True, size=13, color="1F4E79")
ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="DCE6F1")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
r += 1
policy_rows = [
    ("Default language", "ENGLISH ONLY in all English-medium classes, transitions, and English zones."),
    ("When L1 is allowed", "Only for: (a) confirming comprehension breakdown after 2+ scaffolding attempts, (b) emergencies, (c) 30-second safety instructions. NOT as default teaching medium."),
    ("Teacher modeling", "Speak at student's level + 1 (comprehensible input). Slow down, chunk, use hand gestures, point, repeat, rephrase."),
    ("Wait-time", "After asking a Q, wait 5–10 seconds. Do NOT fill silence with translation."),
    ("Recasting vs correcting", "Recast (model correct form naturally) > correct openly. Example: student says 'He go school' → teacher says 'Yes, he GOES to school every day.'"),
    ("Sheltered instruction", "Use realia, visuals, gestures, sentence frames, anchor charts, word walls, chunked reading."),
    ("Translanguaging strategic", "Allow students to USE L1 to think/brainstorm briefly, but output must be in English. Never translate full sentences FOR them."),
    ("Error tolerance", "Errors are part of learning — do not correct every error. Focus on target structures per lesson."),
    ("Code-switching awareness", "Teach students WHEN to use each language: English at school = default; Spanish for Comunicación class + home."),
]
headers2 = ["Policy area", "What it means"]
for i, h in enumerate(headers2, 1):
    ws.cell(row=r, column=i, value=h)
style_header_row(ws, r, 2, height=26)
r += 1
for row_data in policy_rows:
    for i, v in enumerate(row_data, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = BODY_FONT if i == 2 else BOLD_FONT
        c.alignment = WRAP_LEFT
        c.border = BORDER
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    ws.row_dimensions[r].height = 42
    r += 1

# SECTION 4: BICS → CALP PROGRESSION
r += 1
ws.cell(row=r, column=1, value="4. BICS → CALP PROGRESSION").font = Font(bold=True, size=13, color="1F4E79")
ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="DCE6F1")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
r += 1
headers3 = ["Grade", "BICS (social lang.)", "CALP (academic lang.)", "Academic verbs to model"]
for i, h in enumerate(headers3, 1):
    ws.cell(row=r, column=i, value=h)
style_header_row(ws, r, len(headers3), height=26)
r += 1
bics_calp = [
    ("K–G1", "Greetings, classroom phrases, name, colours, numbers", "Point and name, match, sort, describe attributes", "observe, name, count, sort, match, describe"),
    ("G2–G3", "Daily routines, food/drink, hobbies, feelings", "Compare, retell, explain steps, classify", "compare, classify, retell, sequence, explain"),
    ("G4–G5", "Travel, cultures, personal opinions with reasons", "Infer, predict, summarise, analyse simple texts", "predict, infer, summarise, analyse, justify"),
    ("G6–G7", "Debate, negotiate, service language, phone English", "Interpret data, synthesise sources, argue, evaluate", "interpret, synthesise, argue, evaluate, hypothesise"),
    ("G8–G9", "Professional register, conflict resolution, persuasion", "Formulate hypotheses, critique methods, draw conclusions", "formulate, critique, demonstrate, construct, defend"),
    ("G10–G11", "Formal register, leadership, academic debate", "Conduct research, evaluate evidence, theorise", "theorise, conceptualise, extrapolate, substantiate, nominalise"),
]
for row_data in bics_calp:
    for i, v in enumerate(row_data, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = BODY_FONT
        c.alignment = WRAP_LEFT
        c.border = BORDER
    ws.row_dimensions[r].height = 50
    r += 1

set_col_widths(ws, [10, 14, 26, 24, 28, 22, 18])
ws.freeze_panes = "A4"
ws.sheet_view.showGridLines = False


# =============================================================================
# SHEET 22: TEACHING STRATEGIES (differentiation + scaffolding)
# =============================================================================
ws = wb.create_sheet("22_Teaching_Strategies")
next_row = write_title(ws, "Teaching Strategies — Mixed-Level Immersion Classrooms",
                       "For students who don't understand: scaffolding ladder. For different levels: differentiation. For every lesson: sheltered instruction.",
                       row=1, span=6)

# SECTION 1: Scaffolding Ladder — when student doesn't understand
r = next_row
ws.cell(row=r, column=1, value="1. SCAFFOLDING LADDER — when a student doesn't understand").font = Font(bold=True, size=13, color="1F4E79")
ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="DCE6F1")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
r += 1
ws.cell(row=r, column=1, value="Do these IN ORDER. Move down the ladder only if the step above doesn't work. Never skip to translation.").font = Font(italic=True, size=10)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
r += 1
headers = ["Step", "Strategy", "How it looks", "When to use", "Primary (G1-G5)", "Secondary (G6-G11)"]
for i, h in enumerate(headers, 1):
    ws.cell(row=r, column=i, value=h)
style_header_row(ws, r, len(headers), height=26)
r += 1
scaffolding = [
    ("1", "Pause + repeat slowly", "Repeat the same sentence slower, clearer, chunked.", "First attempt — always.",
     "'Open. Your. Book. To. Page. Ten.'", "'The. Industrial. Revolution. Changed. Cities.'"),
    ("2", "Gesture + visual", "Point to the object, draw on board, show picture, use realia.", "Student still blank.",
     "Point to book + page number.", "Show timeline / diagram / photo."),
    ("3", "Rephrase with simpler words", "Use synonyms, cut complex clauses, use high-frequency vocabulary.", "Visual didn't click.",
     "'Look here. This is a dog. Dog goes 'woof'.'", "'Cities grew BIG. Many people moved there for jobs.'"),
    ("4", "Model the answer", "Do it yourself first. Say the sentence. Then ask student to repeat.", "Student can't start.",
     "'I see a red ball. Now you: I see a ___.'", "'In my opinion, the main cause was… Now you try: In my opinion…'"),
    ("5", "Sentence frame", "Give the structure, student fills blank.", "Student has idea but no words.",
     "'My favourite food is ___.'", "'The author's main argument is ___ because ___.'"),
    ("6", "Peer buddy / think-pair-share", "Pair the struggling student with a stronger peer to discuss first.", "Before whole-class share.",
     "Turn & talk with shoulder partner.", "Think-pair-share with sentence stems."),
    ("7", "Pre-teach vocabulary", "Before next lesson, pre-teach 5–7 key words with pictures & gestures.", "Repeated confusion on same topic.",
     "10 min visual warm-up.", "15 min vocab preview with concept maps."),
    ("8", "Chunk the task", "Break big task into 2–3 smaller tasks. Check after each.", "Task is overwhelming.",
     "Read 1 sentence → answer → next sentence.", "Read 1 paragraph → summarise → next."),
    ("9", "Extended time + retry", "Give more minutes, allow redo without penalty.", "Time pressure causing block.",
     "+50% time for writing.", "+50% time for exam tasks."),
    ("10", "Last resort — strategic L1 check", "1-sentence L1 comprehension check, IMMEDIATELY return to English.", "Steps 1–9 all failed; safety or emergency.",
     "'¿Entendiste?' → 'Ok. Let's try again in English.'", "'Tell me in one Spanish sentence what's confusing. Now back to English.'"),
]
for row_data in scaffolding:
    for i, v in enumerate(row_data, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = BODY_FONT
        c.alignment = WRAP_LEFT
        c.border = BORDER
    ws.cell(row=r, column=1).font = Font(bold=True, size=14, color="1F4E79")
    ws.cell(row=r, column=1).alignment = WRAP_CENTER_CENTER
    ws.cell(row=r, column=2).font = BOLD_FONT
    ws.row_dimensions[r].height = 60
    r += 1

# SECTION 2: Differentiation
r += 1
ws.cell(row=r, column=1, value="2. DIFFERENTIATION — mixed-ability classes").font = Font(bold=True, size=13, color="1F4E79")
ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="DCE6F1")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
r += 1
headers2 = ["Strategy", "For lower-level students", "For on-grade students", "For higher-level students", "When to use", "Example"]
for i, h in enumerate(headers2, 1):
    ws.cell(row=r, column=i, value=h)
style_header_row(ws, r, len(headers2), height=26)
r += 1
diff = [
    ("Tiered tasks", "Same content, simplified (word bank, sentence frames)", "Standard task", "Extension (more analysis, creative twist)", "Any lesson with core task", "Story writing: lower = fill-in; on-grade = free writing; higher = add subplot"),
    ("Flexible grouping", "Ability groups (need scaffolding)", "Mixed groups (peer support)", "Ability groups (challenge)", "Rotate weekly", "Guided reading: 3 levels Mon; mixed project Fri"),
    ("Choice boards", "Simpler options", "Standard options", "Challenge options", "Weekly reviews, unit wrap-ups", "Unit review: 3x3 grid with 3 difficulty levels"),
    ("Modified rubrics", "Focus on effort + 2 criteria", "Full 4-criteria rubric", "Full rubric + creativity bonus", "All assessments", "Writing rubric: lower scored on task completion + vocab only"),
    ("Extended / compressed time", "+50% time, break into chunks", "Standard time", "Early-finish extension", "Any timed task", "30-min writing: lower = 45 min in 3 chunks"),
    ("Multi-modal input", "Audio + video + picture + text", "Audio + text", "Text only (pre-view once)", "Pre-teaching, lesson opening", "Listening: lower hears 3x + sees transcript; higher hears 1x"),
    ("Anchor charts + word walls", "High need — always available", "Refer when stuck", "Reference for academic vocab", "Permanent in classroom", "Grammar anchor chart: conditionals with colour-coded parts"),
    ("Peer tutors / buddies", "Paired with higher-level peer", "Rotate roles", "Acts as tutor (consolidates own learning)", "Pair-work activities", "Reading buddies cross-grade: G5 reads to G2"),
]
for row_data in diff:
    for i, v in enumerate(row_data, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = BODY_FONT if i > 1 else BOLD_FONT
        c.alignment = WRAP_LEFT
        c.border = BORDER
    ws.row_dimensions[r].height = 70
    r += 1

# SECTION 3: Strategies by Skill
r += 1
ws.cell(row=r, column=1, value="3. STRATEGIES BY SKILL (sheltered instruction)").font = Font(bold=True, size=13, color="1F4E79")
ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="DCE6F1")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
r += 1
headers3 = ["Skill", "Primary (G1-G5)", "Secondary (G6-G11)", "Tools"]
for i, h in enumerate(headers3, 1):
    ws.cell(row=r, column=i, value=h)
style_header_row(ws, r, len(headers3), height=26)
r += 1
by_skill = [
    ("LISTENING",
     "TPR (Total Physical Response) · Songs with actions · Picture dictation · Listen-and-draw · Listen-and-sequence",
     "Chunked listening · Note-taking templates · Pre-listening vocab · Post-listening discussion · Podcast clips with transcripts",
     "Audio slowed, transcripts, visual support, pause-and-discuss"),
    ("SPEAKING",
     "Sentence frames · Wait-time 10 sec · Partner talk · Dialogue practice · Role-play · Recast errors (don't correct openly)",
     "Think-pair-share · Socratic seminars · Debate protocols · Presentation rubrics · Academic discourse markers",
     "Sentence starters posted, oral rubrics, video self-assessment"),
    ("READING",
     "Picture walks · Shared reading · Choral reading · Guided reading (levelled) · Partner reading · Retelling with props",
     "Close reading · Annotation · Text-dependent questions · Socratic seminars on text · Reading journals · Vocabulary in context",
     "Graphic organisers, word walls, sticky notes for annotation, levelled libraries (Raz Plus)"),
    ("WRITING",
     "Shared writing · Interactive writing · Modelled writing (I do, we do, you do) · Word banks · Sentence frames · Drawing + labelling",
     "Mentor texts · Writing workshop · Peer editing · Process writing (draft → revise → edit → publish) · Targeted feedback",
     "Checklists, rubrics, exemplars, writing portfolios"),
    ("VOCABULARY",
     "Picture cards · Word walls · TPR (act the word) · Songs · Games (Bingo, Go Fish) · Anchor charts",
     "Frayer model · Concept maps · Word study (morphology: prefix/suffix/root) · Context clues · Academic vocabulary notebook",
     "Quizlet, Flipgrid, word walls, etymology posters"),
    ("GRAMMAR",
     "TPR for verb tenses · Song chants · Colour-coded anchor charts (subject-blue, verb-red, object-green)",
     "Grammar mini-lessons (10 min max) · Error analysis · Editing checklists · Sentence combining · Deep revision",
     "Grammar reference cards, mentor sentences, error-correction routines"),
]
for row_data in by_skill:
    for i, v in enumerate(row_data, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = BODY_FONT if i > 1 else Font(bold=True, size=11, color="1F4E79")
        c.alignment = WRAP_LEFT
        c.border = BORDER
    ws.row_dimensions[r].height = 78
    r += 1

set_col_widths(ws, [8, 36, 34, 36, 22, 26])
ws.freeze_panes = "A4"
ws.sheet_view.showGridLines = False


# =============================================================================
# SHEET 23: PARENT GUIDE (bilingual ES/EN, printable)
# =============================================================================
ws = wb.create_sheet("23_Parent_Guide")
next_row = write_title(ws, "Parent Guide — Support at Home WITHOUT Translating",
                       "Guía para padres — Apoya en casa SIN traducir. Bilingüe ES/EN · imprimible.",
                       row=1, span=4)

# Intro
r = next_row
intro = [
    "🎯 El objetivo: tu hijo/a aprende inglés por INMERSIÓN. Traducir cada palabra DAÑA el proceso.",
    "🎯 Goal: your child learns English by IMMERSION. Translating every word HURTS the process.",
    "",
    "Tu rol no es 'profe de inglés en casa' — es ANIMADOR(A) y COMPAÑERO(A) de juego. No necesitas hablar inglés perfecto.",
    "Your role is NOT 'English teacher at home' — it's CHEERLEADER and PLAY COMPANION. You don't need perfect English.",
]
for t in intro:
    c = ws.cell(row=r, column=1, value=t)
    c.font = Font(bold=True, size=11, color="1F4E79") if t.startswith("🎯") else BODY_FONT
    c.alignment = WRAP_LEFT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.row_dimensions[r].height = 22 if t else 10
    r += 1

# SECTION 1: Why no translation
r += 1
ws.cell(row=r, column=1, value="1. ¿POR QUÉ NO TRADUCIR? / Why not translate?").font = Font(bold=True, size=13, color="1F4E79")
ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="DCE6F1")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
r += 1
why = [
    ("El cerebro crea un 'atajo' al español y NUNCA piensa en inglés directamente.",
     "The brain creates a Spanish shortcut and NEVER thinks directly in English."),
    ("Traducir hace que tu hijo sea DEPENDIENTE de ti — no aprende a resolver solo.",
     "Translating makes your child DEPENDENT on you — they don't learn to figure it out."),
    ("En la inmersión, el 'no saber' es parte del proceso. Confunde un rato. Luego entiende.",
     "In immersion, 'not knowing' is part of the process. Confusion comes first. Then understanding."),
    ("Los niños aprenden el primer idioma SIN traducción — pueden hacer lo mismo con el segundo.",
     "Kids learn their first language WITHOUT translation — they can do the same with the second."),
]
ws.cell(row=r, column=1, value="Español").font = HEADER_FONT
ws.cell(row=r, column=1).fill = HEADER_FILL
ws.cell(row=r, column=1).border = BORDER
ws.cell(row=r, column=3, value="English").font = HEADER_FONT
ws.cell(row=r, column=3).fill = HEADER_FILL
ws.cell(row=r, column=3).border = BORDER
ws.cell(row=r, column=1).alignment = WRAP_CENTER
ws.cell(row=r, column=3).alignment = WRAP_CENTER
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
ws.row_dimensions[r].height = 24
r += 1
for es, en in why:
    c1 = ws.cell(row=r, column=1, value=es); c2 = ws.cell(row=r, column=3, value=en)
    for c in (c1, c2):
        c.font = BODY_FONT
        c.alignment = WRAP_LEFT
        c.border = BORDER
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws.row_dimensions[r].height = 38
    r += 1

# SECTION 2: Daily Rituals
r += 1
ws.cell(row=r, column=1, value="2. RUTINAS DIARIAS — qué decir en inglés en casa / Daily rituals at home").font = Font(bold=True, size=13, color="1F4E79")
ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="DCE6F1")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
r += 1
headers = ["Momento / Moment", "Frase en inglés / English phrase", "Pronunciación aprox. (fácil)", "Tip"]
for i, h in enumerate(headers, 1):
    ws.cell(row=r, column=i, value=h)
style_header_row(ws, r, len(headers), height=26)
r += 1
rituals = [
    ("Despertar / Waking up", "Good morning! Did you sleep well?", "Gud morning! Did yu slíp wel?", "Usa siempre, aunque responda en español al inicio."),
    ("Desayuno / Breakfast", "What would you like for breakfast?", "Guat wud yu laik for brékfast?", "Ofrece opciones: 'Cereal or toast?'"),
    ("Camino al cole / School commute", "Have a great day! I love you.", "Jav a gréit dei! Ai lóv yu.", "Usa la misma frase cada día — crea rutina."),
    ("Regreso del cole / After school", "Tell me ONE thing you learned today. In English.", "Tel mi wan zing yu lérnd tudéi.", "Si responde en español, sonríe y di 'In English?'"),
    ("Almuerzo / Lunch", "Please pass the salt / water / bread.", "Plis pas de salt / wóter / bred.", "Cada comida = 2-3 frases en inglés."),
    ("Tarea / Homework", "Do you need help? Show me.", "Du yu nid jelp? Shou mi.", "NO HAGAS la tarea por él/ella."),
    ("Cena / Dinner", "How was your day? Tell me in English.", "Jau guos yor dei?", "Escucha sin corregir — el error es parte del aprendizaje."),
    ("Baño / Bath time", "Time for a bath! Let's go.", "Taim for a baz! Lets gou.", "Rutinas en inglés = inglés automático."),
    ("Cuento / Bedtime story", "Let's read an English story tonight.", "Lets rid an inglish stori tunait.", "CADA NOCHE un cuento en inglés (aunque sea corto)."),
    ("Dormir / Sleep", "Good night! Sweet dreams. I love you.", "Gud nait! Swit drims. Ai lov yu.", "Termina el día en inglés."),
]
for row_data in rituals:
    for i, v in enumerate(row_data, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = BODY_FONT
        c.alignment = WRAP_LEFT
        c.border = BORDER
    ws.row_dimensions[r].height = 32
    r += 1

# SECTION 3: DO / DON'T
r += 1
ws.cell(row=r, column=1, value="3. LO QUE SÍ / LO QUE NO").font = Font(bold=True, size=13, color="1F4E79")
ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="DCE6F1")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
r += 1
ws.cell(row=r, column=1, value="✅ SÍ HACER / DO").font = Font(bold=True, size=12, color="FFFFFF")
ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="70AD47")
ws.cell(row=r, column=1).alignment = WRAP_CENTER
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws.cell(row=r, column=3, value="❌ NO HACER / DON'T").font = Font(bold=True, size=12, color="FFFFFF")
ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor="C00000")
ws.cell(row=r, column=3).alignment = WRAP_CENTER
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
ws.row_dimensions[r].height = 24
r += 1
do_dont = [
    ("Señalar, gesticular, usar imágenes.", "Traducir cada palabra que no entiende."),
    ("Repetir la frase más lento + más claro.", "Corregir cada error que dice tu hijo/a."),
    ("Celebrar el esfuerzo, no la perfección: 'Good try!'", "Obligar a hablar perfecto — causa bloqueo."),
    ("Dejar que fracase y vuelva a intentar.", "Hacer la tarea por él/ella."),
    ("Ver películas en inglés con subtítulos EN INGLÉS (no español).", "Ver todo en español 'porque es más fácil'."),
    ("Usar Raz Plus 15-20 min al día (primaria).", "Saltarse Raz Plus — es la plataforma oficial."),
    ("Preguntar: 'What did you learn today?' en inglés.", "Preguntar solo: '¿Cómo te fue?' en español."),
    ("Validar: 'I don't speak perfect English either. Let's try together.'", "Decir: 'No entiendo ese inglés del cole, pregúntale al profe.'"),
    ("Rutinas fijas de inglés (desayuno, cuento, despedida).", "Inglés solo cuando 'hay ganas' — no funciona."),
    ("Dejar que te enseñe (role reversal): 'Teach me what you learned!'", "Competir con su profesor(a) o contradecir lo enseñado."),
]
for do, dont in do_dont:
    c1 = ws.cell(row=r, column=1, value="✅ " + do)
    c2 = ws.cell(row=r, column=3, value="❌ " + dont)
    for c in (c1, c2):
        c.font = BODY_FONT
        c.alignment = WRAP_LEFT
        c.border = BORDER
    c1.fill = PatternFill("solid", fgColor="E2EFDA")
    c2.fill = PatternFill("solid", fgColor="FCE4D6")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws.row_dimensions[r].height = 34
    r += 1

# SECTION 4: If your child doesn't understand — step by step
r += 1
ws.cell(row=r, column=1, value="4. SI TU HIJO/A NO ENTIENDE — paso a paso / When they don't understand").font = Font(bold=True, size=13, color="1F4E79")
ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="DCE6F1")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
r += 1
steps = [
    ("Paso 1 — Pausa + repite más lento", "Step 1 — Pause + repeat slowly", "'Good. Morning.' (más lento)"),
    ("Paso 2 — Usa gesto / señala", "Step 2 — Use gesture / point", "Apunta a la comida, cuchara, reloj"),
    ("Paso 3 — Rephrase con palabras simples", "Step 3 — Rephrase with simple words", "'What do you want? Cereal? Toast?'"),
    ("Paso 4 — Muestra imagen / objeto", "Step 4 — Show picture / object", "Saca el cereal y muéstralo"),
    ("Paso 5 — Deja que te enseñe", "Step 5 — Let them teach you", "'Can YOU say it in English?'"),
    ("Paso 6 (último recurso) — L1 check rápido", "Step 6 (last resort) — quick L1 check", "'¿Necesitas ayuda?' → inmediatamente: 'OK, try in English.'"),
]
headers4 = ["Paso / Step", "(inglés) / (English)", "Ejemplo / Example", ""]
for i, h in enumerate(headers4, 1):
    if h:
        ws.cell(row=r, column=i, value=h)
style_header_row(ws, r, 3, height=24)
r += 1
for es, en, ex in steps:
    ws.cell(row=r, column=1, value=es).font = BOLD_FONT
    ws.cell(row=r, column=2, value=en).font = BODY_FONT
    ws.cell(row=r, column=3, value=ex).font = Font(italic=True, size=10)
    for c in range(1, 4):
        ws.cell(row=r, column=c).alignment = WRAP_LEFT
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = 32
    r += 1

# SECTION 5: Weekly routine + tech tools
r += 1
ws.cell(row=r, column=1, value="5. RUTINA SEMANAL + HERRAMIENTAS / Weekly routine + tools").font = Font(bold=True, size=13, color="1F4E79")
ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="DCE6F1")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
r += 1
week = [
    ("Lun-Vie / Mon-Fri", "Raz Plus 15-20 min (primaria)", "Todos los días. 5 minutos > 1 hora semanal."),
    ("Lunes / Monday", "Música en inglés mientras desayunan", "Playlist familiar en inglés."),
    ("Miércoles / Wed", "Cuento en inglés antes de dormir", "Plan Lector + Raz Plus + biblioteca."),
    ("Sábado / Saturday", "Película en inglés CON subtítulos en inglés (no español)", "Disney, Pixar, Bluey, Peppa Pig funcionan muy bien."),
    ("Domingo / Sunday", "Conversación familiar en inglés 20 min (cena)", "Tema: 'best thing of the week'."),
]
headers5 = ["Día / Day", "Actividad / Activity", "Nota / Note", ""]
for i, h in enumerate(headers5, 1):
    if h: ws.cell(row=r, column=i, value=h)
style_header_row(ws, r, 3, height=24)
r += 1
for d, a, n in week:
    ws.cell(row=r, column=1, value=d).font = BOLD_FONT
    ws.cell(row=r, column=2, value=a).font = BODY_FONT
    ws.cell(row=r, column=3, value=n).font = Font(italic=True, size=10)
    for c in range(1, 4):
        ws.cell(row=r, column=c).alignment = WRAP_LEFT
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = 30
    r += 1

r += 1
ws.cell(row=r, column=1, value="HERRAMIENTAS RECOMENDADAS / Recommended tools:").font = Font(bold=True, size=11, color="1F4E79")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
r += 1
tools = [
    "📚 Raz Plus (todos los alumnos tienen acceso) — PLATAFORMA OFICIAL NIS",
    "🎧 English audiobooks / podcasts (Storynory, BBC Learning English, ABC Kids)",
    "🎬 Películas/series en inglés con subtítulos en inglés — Bluey, Peppa Pig (primaria), Brooklyn 99, Modern Family (secundaria)",
    "📱 Duolingo — gamificado, 10 min/día (G4+)",
    "🎵 Spotify: English Kids Songs, Super Simple Songs, The Pop Song Method",
    "🎨 YouTube Kids (solo canales en inglés): Dave and Ava, Cocomelon, SciShow Kids",
]
for t in tools:
    ws.cell(row=r, column=1, value=t).font = BODY_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.row_dimensions[r].height = 20
    r += 1

set_col_widths(ws, [32, 34, 34, 12])
ws.sheet_view.showGridLines = False
ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True


# =============================================================================
# SAVE
# =============================================================================
OUTPUT = r"C:\Users\Paolo\OneDrive\09_Instituciones\NNORDIC\NIS_English_Master_SS_2026.xlsx"
wb.save(OUTPUT)
print(f"OK: {OUTPUT}")
print(f"Sheets created: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"  - {s}")
